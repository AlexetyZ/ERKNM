import openpyxl
import re
class Table_gasu:
    def __init__(self):
        wb_path = "C:\\Users\\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\\2022.xlsx"
        self.template_path = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\форма госдоклады - шаблон не перемещать.xlsx"
        self.report_book = openpyxl.load_workbook(self.template_path)
        self.sh_template = self.report_book.worksheets[0]

        self.wb = openpyxl.load_workbook(wb_path)
        self.sh_kind_controll_1 = self.wb.worksheets[0]
        self.sh_kind_activity_2 = self.wb.worksheets[1]
        self.sh_organization_3 = self.wb.worksheets[2]
        self.sh_prof_organization_4 = self.wb.worksheets[3]
        self.sh_detail_all_5 = self.wb.worksheets[4]
        self.sh_knm_detail_6 = self.wb.worksheets[5]

    def driver(self):
        # self.count_knm_and_breaches('потреб')
        locus_dict = {
            'анитарно-эпидем': 6,
            'иониз': 5,
            'возбуд': 4,
            'детей': 3,
            'потреб': 2,
        }
        for controll_kind, controll_locus in locus_dict.items():
            self.count_knm_and_breaches(kind_controll=controll_kind, locus_controll=controll_locus)

    def count_knm_and_breaches(self, kind_controll, locus_controll: int):
        """
        Поиск по таблице ГАСУ количество проведенных кнм с взяимодействием
        @param locus_controll: номер столбца в шаблоне отчета, который соответствует виду контроля
        @param kind_controll: вводим вид контроля - проверка на содержание отрезка, например если в значении содержится ..., то... ех: "Маршрутка" содрежит "Марш"
        @return:
        """

        controll_plan = 0
        controll_plan_purchase = 0
        controll_plan_purchase_msp = 0
        controll_plan_monitoring = 0
        controll_plan_monitoring_msp = 0
        controll_plan_selective = 0
        controll_plan_selective_msp = 0
        controll_plan_inspect_visit = 0
        controll_plan_inspect_visit_msp = 0
        controll_plan_reid = 0
        controll_plan_reid_msp = 0
        controll_plan_doc = 0
        controll_plan_doc_msp = 0
        controll_plan_onsite = 0
        controll_plan_onsite_msp = 0

        controll_outplan = 0
        controll_outplan_purchase = 0
        controll_outplan_purchase_msp = 0
        controll_outplan_monitoring = 0
        controll_outplan_monitoring_msp = 0
        controll_outplan_selective = 0
        controll_outplan_selective_msp = 0
        controll_outplan_inspect_visit = 0
        controll_outplan_inspect_visit_msp = 0
        controll_outplan_reid = 0
        controll_outplan_reid_msp = 0
        controll_outplan_doc = 0
        controll_outplan_doc_msp = 0
        controll_outplan_onsite = 0
        controll_outplan_onsite_msp = 0

        count_breaches = 0
        count_breaches_msp = 0

        controlls_with_breach = 0
        controlls_with_breach_msp = 0

        count_prescription_breach = 0


        for row in self.sh_knm_detail_6.iter_rows(4, values_only=True):
            if kind_controll in row[1]:
                count_breaches += int(row[8])
                count_breaches_msp += int(row[12])
                if 'предпис' in str(row[4]):
                    count_prescription_breach += int(row[8])
                controlls_with_breach += int(row[7])
                controlls_with_breach_msp += int(row[11])
                if row[2] == 'Плановая проверка':
                    controll_plan += int(row[6])
                    if 'выездная' in str(row[3]).lower():
                        controll_plan_onsite += int(row[6])
                        controll_plan_onsite_msp += int(row[10])
                        continue
                    if 'документар' in str(row[3]).lower():
                        controll_plan_doc += int(row[6])
                        controll_plan_doc_msp += int(row[10])
                        continue
                    if 'контрольна' in str(row[3]).lower():
                        controll_plan_purchase += int(row[6])
                        controll_plan_purchase_msp += int(row[10])
                        continue
                    if 'мониторинг' in str(row[3]).lower():
                        controll_plan_monitoring += int(row[6])
                        controll_plan_monitoring_msp += int(row[10])
                        continue
                    if 'выборочн' in str(row[3]).lower():
                        controll_plan_selective += int(row[6])
                        controll_plan_selective_msp += int(row[10])
                        continue
                    if 'инспекц' in str(row[3]).lower():
                        controll_plan_inspect_visit += int(row[6])
                        controll_plan_inspect_visit_msp += int(row[10])
                        continue
                    if 'рейд' in str(row[3]).lower():
                        controll_plan_reid += int(row[6])
                        controll_plan_reid_msp += int(row[10])
                        continue

                if row[2] == 'Внеплановая проверка':
                    controll_outplan += int(row[6])
                    if 'выездная' in str(row[3]).lower():
                        controll_outplan_onsite += int(row[6])
                        controll_outplan_onsite_msp += int(row[10])
                        continue
                    if 'документар' in str(row[3]).lower():
                        controll_outplan_doc += int(row[6])
                        controll_outplan_doc_msp += int(row[10])
                        continue
                    if 'контрольна' in str(row[3]).lower():
                        controll_outplan_purchase += int(row[6])
                        controll_outplan_purchase_msp += int(row[10])
                        continue
                    if 'мониторинг' in str(row[3]).lower():
                        controll_outplan_monitoring += int(row[6])
                        controll_outplan_monitoring_msp += int(row[10])
                        continue
                    if 'выборочн' in str(row[3]).lower():
                        controll_outplan_selective += int(row[6])
                        controll_outplan_selective_msp += int(row[10])
                        continue
                    if 'инспекц' in str(row[3]).lower():
                        controll_outplan_inspect_visit += int(row[6])
                        controll_outplan_inspect_visit_msp += int(row[10])
                        continue
                    if 'рейд' in str(row[3]).lower():
                        controll_outplan_reid += int(row[6])
                        controll_outplan_reid_msp += int(row[10])
                        continue
        summ_plan = controll_plan_purchase + controll_plan_monitoring + controll_plan_selective + controll_plan_inspect_visit + controll_plan_reid + controll_plan_doc + controll_plan_onsite
        if controll_plan == summ_plan:
            print('Проверка по плановым прошла успешно - все сошлось!')
        else:
            print(f'!!!Не сошлись цифры плановая проверка: plan-{controll_plan}, summ - {summ_plan}')


        self.sh_template.cell(row=15, column=locus_controll, value=controll_plan)
        self.sh_template.cell(row=16, column=locus_controll, value=controll_plan_purchase)
        self.sh_template.cell(row=17, column=locus_controll, value=controll_plan_purchase_msp)
        self.sh_template.cell(row=18, column=locus_controll, value=controll_plan_monitoring)
        self.sh_template.cell(row=19, column=locus_controll, value=controll_plan_monitoring_msp)
        self.sh_template.cell(row=20, column=locus_controll, value=controll_plan_selective)
        self.sh_template.cell(row=21, column=locus_controll, value=controll_plan_selective_msp)
        self.sh_template.cell(row=22, column=locus_controll, value=controll_plan_inspect_visit)
        self.sh_template.cell(row=23, column=locus_controll, value=controll_plan_inspect_visit_msp)
        self.sh_template.cell(row=24, column=locus_controll, value=controll_plan_reid)
        self.sh_template.cell(row=25, column=locus_controll, value=controll_plan_reid_msp)
        self.sh_template.cell(row=26, column=locus_controll, value=controll_plan_doc)
        self.sh_template.cell(row=27, column=locus_controll, value=controll_plan_doc_msp)
        self.sh_template.cell(row=28, column=locus_controll, value=controll_plan_onsite)
        self.sh_template.cell(row=29, column=locus_controll, value=controll_plan_onsite_msp)

        self.sh_template.cell(row=30, column=locus_controll, value=controll_outplan)
        self.sh_template.cell(row=31, column=locus_controll, value=controll_outplan_purchase)
        self.sh_template.cell(row=32, column=locus_controll, value=controll_outplan_purchase_msp)
        self.sh_template.cell(row=33, column=locus_controll, value=controll_outplan_monitoring)
        self.sh_template.cell(row=34, column=locus_controll, value=controll_outplan_monitoring_msp)
        self.sh_template.cell(row=35, column=locus_controll, value=controll_outplan_selective)
        self.sh_template.cell(row=36, column=locus_controll, value=controll_outplan_selective_msp)
        self.sh_template.cell(row=37, column=locus_controll, value=controll_outplan_inspect_visit)
        self.sh_template.cell(row=38, column=locus_controll, value=controll_outplan_inspect_visit_msp)
        self.sh_template.cell(row=39, column=locus_controll, value=controll_outplan_reid)
        self.sh_template.cell(row=40, column=locus_controll, value=controll_outplan_reid_msp)
        self.sh_template.cell(row=41, column=locus_controll, value=controll_outplan_doc)
        self.sh_template.cell(row=42, column=locus_controll, value=controll_outplan_doc_msp)
        self.sh_template.cell(row=43, column=locus_controll, value=controll_outplan_onsite)
        self.sh_template.cell(row=44, column=locus_controll, value=controll_outplan_onsite_msp)

        self.sh_template.cell(row=75, column=locus_controll, value=count_breaches)
        self.sh_template.cell(row=76, column=locus_controll, value=count_breaches)
        self.sh_template.cell(row=77, column=locus_controll, value=count_breaches_msp)

        self.sh_template.cell(row=79, column=locus_controll, value=controlls_with_breach)
        self.sh_template.cell(row=80, column=locus_controll, value=controlls_with_breach_msp)

        self.sh_template.cell(row=81, column=locus_controll, value=count_prescription_breach)



        self.report_book.save(self.template_path)



        print(f'{controll_plan=}')
        print(f'{controll_plan_purchase=}')
        print(f'{controll_plan_purchase_msp=}')
        print(f'{controll_plan_monitoring=}')
        print(f'{controll_plan_monitoring_msp=}')
        print(f'{controll_plan_selective=}')
        print(f'{controll_plan_selective_msp=}')
        print(f'{controll_plan_inspect_visit=}')
        print(f'{controll_plan_inspect_visit_msp=}')
        print(f'{controll_plan_reid=}')
        print(f'{controll_plan_reid_msp=}')
        print(f'{controll_plan_doc=}')
        print(f'{controll_plan_doc_msp=}')
        print(f'{controll_plan_onsite=}')
        print(f'{controll_plan_onsite_msp=}')

        summ_outplan = controll_outplan_purchase + controll_outplan_monitoring + controll_outplan_selective + controll_outplan_inspect_visit + controll_outplan_reid + controll_outplan_doc + controll_outplan_onsite
        if controll_outplan == summ_outplan:
            print('Проверка по внеплановым прошла успешно - все сошлось!')
        else:
            print(f'!!!Не сошлись цифры внеплановая проверка: outplan-{controll_plan}, summ - {summ_plan}')

        print(f'{controll_outplan=}')
        print(f'{controll_outplan_purchase=}')
        print(f'{controll_outplan_purchase_msp=}')
        print(f'{controll_outplan_monitoring=}')
        print(f'{controll_outplan_monitoring_msp=}')
        print(f'{controll_outplan_selective=}')
        print(f'{controll_outplan_selective_msp=}')
        print(f'{controll_outplan_inspect_visit=}')
        print(f'{controll_outplan_inspect_visit_msp=}')
        print(f'{controll_outplan_reid=}')
        print(f'{controll_outplan_reid_msp=}')
        print(f'{controll_outplan_doc=}')
        print(f'{controll_outplan_doc_msp=}')
        print(f'{controll_outplan_onsite=}')
        print(f'{controll_outplan_onsite_msp=}')

        print('')
        print('')
        print(f'{count_breaches=}')
        print(f'{count_breaches_msp=}')
        print('')
        print('')
        print(f'{controlls_with_breach=}')
        print(f'{controlls_with_breach_msp=}')


    def count_ul_and_breaches(self, kind_controll):
        all_inns_have_breaches = []
        ul_inns_have_breaches = []
        for row in self.sh_organization_3.iter_rows(3, values_only=True):
            if kind_controll in row[2]:
                all_inns_have_breaches.append(row[1])
                if row[4]:
                    ul_inns_have_breaches.append(row[1])




if __name__ == '__main__':

    Table_gasu().driver()
    # Table_gasu().