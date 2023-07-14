from EGRUL import Egrul
import openpyxl
from tqdm import tqdm
import time

class Find:
    def __init__(self):
        self.e = Egrul()

    def from_xl(self, path: str, x: int = 23, y: int = 24):
        wb = openpyxl.load_workbook(path)
        sh = wb.worksheets[0]
        for row in tqdm(range(sh.max_row+1)):
            inn = sh.cell(row=row+1, column=x).value
            # print(inn)
            try:
                request = self.e.find_in_egrul(inn)
                sh.cell(row=row + 1, column=y, value=request['r'])
                time.sleep(1)
            except Exception as ex:
                wb.save(path)
                raise ValueError(f'Ошибка запроса: {ex}')
        wb.save(path)


if __name__ == '__main__':
    Find().from_xl("S:\\Зайцев_АД\Выверка ЕРВК\заносить вручную\то шо у плане2.xlsx")


