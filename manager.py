import os
import sys
import openpyxl


def main(filePath):
    wb = openpyxl.load_workbook(filePath)
    sh1 = wb.worksheets[1]
    for col in sh1.iter_cols():
        print(col[0].is_merged)
        print(sh1.column_dimensions[col[0].column_letter].width)
    # for row in sh1.iter_rows():
    #     print([cell.style for cell in row])


if __name__ == '__main__':
    filePath = '/Users/aleksejzajcev/Documents/отчеты/План 2024/Ежедневный отчет о ходе согласования/ежедневный 06.11.2023 knm.xlsx'
    main(filePath)


