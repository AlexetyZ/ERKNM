import openpyxl
import os


def main():
    wb_template = openpyxl.load_workbook('шаблон/!Шаблон.xlsx')
    sh_template = wb_template.worksheets[0]
    for file in os.listdir('сырые файлики'):
        wb_file = openpyxl.load_workbook(f'сырые файлики/{file}')
        sh_file = wb_file.worksheets[1]
        for row in range(10, 30):
            value = sh_file.cell(row=row, column=7).value
            if not value:
                value = 0
            sh_template.cell(row=row-7, column=5, value=value)
        wb_template.save(f'готовые файлики/{file}')


if __name__ == '__main__':
    main()
