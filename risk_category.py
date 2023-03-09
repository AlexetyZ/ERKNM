import openpyxl
import os
from sql import Database
from datetime import datetime
print('Начало программы')
wb_path = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\перечень объектов контроля\Виды деятельности_на сайт_25012023.xlsx"
wb = openpyxl.load_workbook(wb_path)
sh = wb.worksheets[0]
d = Database()
sh.cell(row=1, column=7, value='1 - субъект есть в плане, но объекта с такой категорией риска нет')
print(f"{datetime.now()} файл прочитан, начинаем проверку...")


def main():
    for n, row in enumerate(sh.iter_rows(min_row=3)):
        if n % 10000 == 0:
            print(f'{datetime.now()} - {n}')
        inn = row[3].value
        object_info = d.is_subject_exists(inn)

        risk = str(row[5].value).lower()
        # print(risk)

        if object_info:
            obj_risk = []
            exists_risk_category = False
            for obj in object_info:
                o_risk = obj[1]
                if o_risk in risk:
                    exists_risk_category = True
            if not exists_risk_category:
                sh.cell(row=n+3, column=7, value='1')

    wb.save(wb_path)

def checks_exists_subject():
    inn = '9102188171'
    exists_subject = d.is_subject_exists(inn)
    for row in exists_subject:
        print(row)


if __name__ == '__main__':
    main()
    # checks_exists_subject()
