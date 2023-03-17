import openpyxl
from pathlib import Path
from openpyxl.comments import Comment



class Check:
    def __init__(self, wb_path_1):
        self.wb_path_1 = wb_path_1
        self.wb_1 = openpyxl.load_workbook(self.wb_path_1)

    def main(self, *sheet_index):
        min_row = 3
        min_col = 2
        text_errors = f'Ошибки в файле {self.wb_path_1}'
        for index in sheet_index:

            check_logic = self.logic(sheet_index=index, min_row=min_row, min_col=min_col)
            text_errors += f'\nНа листе {self.wb_1.sheetnames[index]} ({self.wb_1.worksheets[index]["C1"].value}):'
            if check_logic:
                for error in check_logic:
                    text_errors += f'\n\t - столбец {error["column"]}, строка {error["row"]}: {error["error"]}'
                    self.wb_1.worksheets[index].cell(column=error["column"], row=error["row"]).comment = Comment(text=error["error"], author='Зайцев Алексей Дмитриевич')
            else:
                text_errors += f'\n\t - Все верно!'

        self.wb_1.save(self.wb_path_1)
        with open(f'{Path(self.wb_path_1).parent}/error_file {Path(self.wb_path_1).stem}.txt', 'w') as file:
            file.write(text_errors)



    def leave_coment(self, sheet_index, row, column, text):
        sh = self.wb_1.worksheets[sheet_index]
        sh.cell(row=row, column=column, value=text)


    def logic(self, sheet_index: int, min_row: int = 3, min_col: int = 2):
        """
        Функция проверяет сумму слагаемых на верность. Если сумма слагаемых не равна обещему количеству, возвращается список с ошибками
        :param index_main: индекс общего количества из кортежа значений
        :param list_index_parts: список индексов, сумма которых должна быть равна общей сумме
        :param min_row: минимальная строка, задевающая первую строку значений и остаток шапки (последнюю строку шапки)
        :param min_col: минимальная колонка, затрагивающая только численные значения строк
        :return: Список ошибок. при отсутствии возвращает пустой список
        """
        errors = []
        sh_0 = self.wb_1.worksheets[sheet_index]
        # print(len(self.wb_1.worksheets))



        # проверка логики по столбцам
        for col_number, column in enumerate(sh_0.iter_cols(min_col=min_col, min_row=min_row, values_only=True)):
            def check_sum(all_index: int, list_parts_indexes: list):
                """
                Проверяем общее число и сумму слагаемых
                :param all_index:
                :param list_parts_indexes:
                :return:
                Ничего не возвращает, просто наполняет журнал ошибок значениями
                """
                summ = self._is_summ_errors(column, all_index, list_parts_indexes)
                if summ:
                    errors.append({'column': min_col + col_number, 'row': min_row + all_index,
                                   'error': f'{summ} (стр. {", ".join(str(lpi + min_row) for lpi in list_parts_indexes)})'})

            def check_exceed(all_index: int, list_parts_indexes: list):
                amount_error = self._is_exseed_amount_error(column, all_index, list_parts_indexes)
                if amount_error:
                    errors.append({'column': min_col + col_number, 'row': min_row + all_index,
                                   'error': f'{amount_error} (стр. {", ".join(str(lpi + min_row) for lpi in list_parts_indexes)})'})

            check_sum(1, [2, 3, 4, 5, 6, 7, 9])
            check_sum(11, [12, 27])
            check_sum(12, [13, 15, 17, 19, 21, 23, 25])
            check_sum(27, [28, 30, 32, 34, 36, 38, 40])
            check_sum(42, [43, 44, 45, 46, 47, 49, 50, 51, 52, 53])
            check_sum(56, [57, 58, 59])
            check_sum(62, [63, 64, 65])
            check_sum(68, [69, 70, 71])
            check_sum(83, [84, 85, 86, 87, 88, 89, 90, 91])
            check_sum(91, [92, 93, 94, 95])
            check_sum(96, [97, 98, 99, 100])
            check_sum(102, [103, 104])
            check_sum(107, [108, 110])
            check_sum(112, [113, 114, 115])

            check_exceed(7, [8])
            check_exceed(9, [10])

            check_exceed(13, [14])
            check_exceed(15, [16])
            check_exceed(17, [18])
            check_exceed(19, [20])
            check_exceed(21, [22])
            check_exceed(23, [24])
            check_exceed(25, [26])

            check_exceed(28, [29])
            check_exceed(30, [31])
            check_exceed(32, [33])
            check_exceed(34, [35])
            check_exceed(36, [37])
            check_exceed(38, [39])
            check_exceed(40, [41])

            check_exceed(47, [48])
            check_exceed(60, [61])
            check_exceed(66, [67])

            check_exceed(73, [74])
            check_exceed(76, [77])

            check_exceed(81, [82])

            check_exceed(108, [109])
            check_exceed(110, [111])
            check_exceed(119, [120])
            check_exceed(121, [122])
            check_exceed(123, [124])
            check_exceed(126, [127])
            check_exceed(128, [129])


        return errors

    def pretrial_appeal_in_tu(self, *sheet_index):
        columns = ['F']
        summ_value = 0
        for column in columns:
            for index in sheet_index:
                sheet = self.wb_1.worksheets[index]

                value = sheet[f'{column}117'].value
                if value:
                    try:
                        summ_value += value
                    except Exception as ex:
                        print(ex)
                    print(f"{sheet['C1'].value} | {value} | ({sheet[f'{column}5'].value})")
        print('')
        print(f'Общая сумма - {summ_value}')



    def _is_summ_errors(self, column, all_index: int, list_parts_indexes: list):
        """
        Считает данные общей величины и суммы слагаемых и возвращает, если не совпадает
        :param column: кортеж из значений
        :param all_index: индекс общей величины в кортеже из значений (column)
        :param list_parts_indexes: индексы слагаемых в кортеже из значений (column)
        :return: если несовпадение, возвращает ошибку строкой.
        """
        all_m_count = column[all_index]
        if isinstance(all_m_count, (int, float)):
            sum_m_count = sum([column[i] for i in list_parts_indexes if isinstance(column[i], (int, float))])
            if all_m_count != sum_m_count:
                return f'ошибка суммы: {all_m_count} не равно {sum_m_count}'

    def _is_exseed_amount_error(self, column, all_index: int, list_parts_indexes: list):
        all_m_count = column[all_index]
        if isinstance(all_m_count, (int, float)):
            sum_m_count = sum([column[i] for i in list_parts_indexes if isinstance(column[i], (int, float))])
            if all_m_count < sum_m_count:
                return f'ошибка превышения суммы: {all_m_count} должно быть больше {sum_m_count}'


if __name__ == '__main__':
    aggree = input('Рекомендуется сделать резервную копию рабочего файла на случай утери. если Вы все поняли и согласны с риском, введите "д"')
    if aggree.lower() == 'д' or aggree.lower() == 'l':
        # gd_path = '/Volumes/KINGSTON/гасу/ФБУЗ/Исходные от субъектов РФ_Доклад 2022 -10-03-2023.xlsx'
        # gd_path = '/Volumes/KINGSTON/гасу/ФБУЗ/Форма своданя госдоклады_09.03.2023_ (1).xlsx'

        # gd_path = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\Окончательная Форма своданя госдоклады_15.03.2023 по 1-22___.xlsx"
        gd_path = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\ФБУЗ\новая свод по ТУ.xlsx"

        # print(iter(1,2)
        # Check(gd_path).main(0)
        Check(gd_path).pretrial_appeal_in_tu(*(l for l in range(0, 86)))


