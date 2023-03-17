from main_ERKNM import erknm
import openpyxl
from data_description import rename_tu_dict

class Table_erknm:
    def __init__(self, wb_path_1, wb_path_2: str = ''):
        self.wb_path_1 = wb_path_1
        self.wb_1 = openpyxl.load_workbook(self.wb_path_1)
        if wb_path_2:
            self.wb_path_2 = wb_path_2
            self.wb_2 = openpyxl.load_workbook(self.wb_path_2)

    def rename_tu_list(self, sh_index: int = 0):
        sh = self.wb_1.worksheets[sh_index]
        for n, row in enumerate(sh.iter_rows(min_row=4, values_only=True)):
            try:
                new_tu_name = rename_tu_dict[row[1]]
            except:
                continue
            sh.cell(column=2, row=n+4, value=new_tu_name)
        self.wb_1.save(self.wb_path_1)

    def svod_gasu_erknm_meri(self):
        sh_erknm = self.wb_2.worksheets[0]
        sh_gasu = self.wb_1.worksheets[5]

        for n, row_erknm in enumerate(sh_erknm.iter_rows(min_row=2, values_only=True)):
            tu_plan_count = 0
            tu_outplan_count = 0
            tu_name_erknm = row_erknm[1]
            for row_gasu in sh_gasu.iter_rows(min_row=4, values_only=True):
                if row_gasu[0] == tu_name_erknm:
                    if row_gasu[2] == 'Плановая проверка':
                        tu_plan_count += row_gasu[6]
                    else:
                        tu_outplan_count += row_gasu[6]
            sh_erknm.cell(column=5, row=n+2, value=tu_plan_count)
            sh_erknm.cell(column=6, row=n+2, value=tu_outplan_count)
        self.wb_2.save(self.wb_path_2)



    def compare_regions_doc_and_erknm(self):
        """
        region docs = wb_1
        erknm = wb_2
        @return:
        """

        for erknm_report_row in self.wb_2.worksheets[0].iter_rows(min_row=3, values_only=True):

            erknm_tu_name = erknm_report_row[1]
            for sheet_in_region_doc in self.wb_1.worksheets:
                if sheet_in_region_doc['C1'] == erknm_tu_name:
                    for cell_g in range(16, 45):
                        sheet_in_region_doc[f'G{cell_g}'].value = erknm_report_row[cell_g-14]

        self.wb_1.save(self.wb_path_1)


    def rename_tu_in_pages(self):
        for page in self.wb_1.worksheets:
            old_tu_name = page['C1'].value
            try:
                new_tu_name = rename_tu_dict[old_tu_name]
            except:
                continue
            page['C1'].value = new_tu_name
        self.wb_1.save(self.wb_path_1)





    def main(self):
        pass

    def create_dict(self):
        """

        @return:
        """
        tu_dict = {}
        sh = self.wb_1.worksheets[0]
        for row in sh.iter_rows(min_row=2, values_only=True):
            tu_dict[row[4]] = row[2]
        return tu_dict


if __name__ == '__main__':
    # path_table_1_1_22 = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\Субъекты РФ_ табл 1.xlsx"
    # path_table_2_1_22 = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\Субъекты РФ_ таблица 2.xlsx"
    # path_gasu = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\\2022.xlsx"
    # path_erknm = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\отчет ЕРКНМ 2023-03-10.xlsx"

    fbuz_regions = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\ФБУЗ\новая свод по ТУ.xlsx"

    report_erknm = "C:\\Users\zaitsev_ad\Documents\ЕРКНМ\отчеты ГАСУ\отчет КНМ.xlsx"

    Table_erknm(
        # path_table_1_1_22,
        # path_table_2_1_22,
        # path_gasu,
        # path_erknm
        fbuz_regions,
        # report_erknm
    ).rename_tu_in_pages()
