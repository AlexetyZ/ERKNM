import time
from tqdm import tqdm
import openpyxl
from sys import argv
import os
from pprint import pprint
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import cell
import re


def formatFile(
        pathFile,
        columsWidth: dict,
        cellsWrap=None,
        mergeCells=None,
        positionText=None,
        formatNumber=None,
        borderForAllCellInRange=None,
        fonts: dict = None,
        paperSize: int = None,
        printArea: str = None,
        printCols: str = None,
        printRows: str = None,
        fitToHeight: int = None,
        fitToWidht: int = None,
        sheetIndex: int = 0,

):
    # wb = openpyxl.load_workbook(pathFile)
    wb = openpyxl.load_workbook(pathFile)
    sh = wb.worksheets[sheetIndex]

    def stabillizeCoordinates(coordinate: str):
        functions = ['max']
        if ':' in coordinate:
            parts = coordinate.split(':')
        else:
            parts = [coordinate]
        newParts = []
        for part in parts:
            if 'max' in part:
                # print(column)
                part = part.replace('max', str(sh.max_row))
            newParts.append(part)
        return ':'.join(newParts)




    for col, width in columsWidth.items():
        sh.column_dimensions[col].width = width
    if mergeCells:
        if isinstance(mergeCells, str):
            mergeCells = [mergeCells]
        for cells in mergeCells:
            cells = stabillizeCoordinates(cells)
            sh.merge_cells(cells)

    if positionText:
        for cells, position in positionText.items():
            cells = stabillizeCoordinates(cells)
            if isinstance(sh[cells], tuple):
                for cell in sh[cells]:
                    if isinstance(cell, tuple):
                        for c in cell:

                            c.alignment += Alignment(horizontal=position, vertical=position)
                    else:

                        cell.alignment += Alignment(horizontal=position, vertical=position)

    if cellsWrap:
        if isinstance(cellsWrap, str):
            cellsWrap = [cellsWrap]
        for cells in cellsWrap:
            cells = stabillizeCoordinates(cells)
            if isinstance(sh[cells], tuple):
                for cell in sh[cells]:
                    if isinstance(cell, tuple):
                        for c in cell:
                            c.alignment += Alignment(wrap_text=True)
                    else:
                        cell.alignment += Alignment(wrap_text=True)

    if formatNumber:
        for cells, format in formatNumber.items():
            cells = stabillizeCoordinates(cells)
            if isinstance(sh[cells], tuple):
                for cell in sh[cells]:
                    if isinstance(cell, tuple):
                        for c in cell:
                            c.number_format = format
                    else:
                        cell.number_format = format

    if borderForAllCellInRange:
        thin = Side(border_style="thin", color="000000")
        for cells in borderForAllCellInRange:
            cells = stabillizeCoordinates(cells)
            if isinstance(sh[cells], tuple):
                for cell in sh[cells]:
                    if isinstance(cell, tuple):
                        for c in cell:
                            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    else:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    if fonts:
        for cells, style in fonts.items():
            cells = stabillizeCoordinates(cells)
            if isinstance(sh[cells], tuple):
                for cell in sh[cells]:
                    if isinstance(cell, tuple):
                        for c in cell:
                            c.font = style
                    else:
                        cell.font = style

    if paperSize:
        sh.page_setup.paperSize = paperSize

    if printArea:
        printArea = stabillizeCoordinates(printArea)
        sh.print_area = printArea
    if printCols:
        sh.print_title_cols = printCols

    if printRows:
        print(printRows)
        sh.print_title_rows = f'{printRows}'

    if fitToHeight:
        if not sh.sheet_properties.pageSetUpPr.fitToPage:
            sh.sheet_properties.pageSetUpPr.fitToPage = True
        sh.page_setup.fitToHeight = fitToHeight

    if fitToWidht:
        if not sh.sheet_properties.pageSetUpPr.fitToPage:
            sh.sheet_properties.pageSetUpPr.fitToPage = True
        sh.page_setup.fitToWidht = fitToWidht


    wb.save(pathFile)



def dellnull(pathDir):
    for file in os.listdir(pathDir):
        pathFile = os.path.join(pathDir, file)
        wb = openpyxl.load_workbook(pathFile)
        sh = wb.worksheets[0]
        if 'Идентификатор ЕИАС' in sh['A2'].value and not sh['A3'].value:
            os.remove(pathFile)


def bringCol(pathFile, minCol: str or int or bool = 1, minRow: str or int or bool = 2,
             colNumber: str or int or bool = None, sheetIndex: int = 0) -> list:
    """
    pathFile: путь до файла эксель
    minCol: минимальный столбец, с которого отчитываем
    minRow: минимальный столбец, с которого отчитываем
    colNumber: столбец, значения из которого возвращаются
    """
    minCol = minCol if minCol and str(minCol).isdigit() is not None else '1'
    minRow = minRow if minRow and str(minRow).isdigit() is not None else '2'
    colNumber = colNumber if colNumber else '0'
    wb = openpyxl.load_workbook(pathFile)
    sh = wb.worksheets[sheetIndex]

    if ':' in str(colNumber):
        se = [int(val) if val else None for val in colNumber.split(':')]
        colNumber = slice(se[0], se[1])
    else:
        try:
            colNumber = int(colNumber)
        except:
            colNumber = 0

    return [row for row in sh.iter_cols(min_col=int(minCol), min_row=int(minRow), values_only=True)][colNumber]


def bringRowsVal(path, sheetIndex, minRow, minCol):
    minCol = minCol if minCol and str(minCol).isdigit() is not None else '1'
    minRow = minRow if minRow and str(minRow).isdigit() is not None else '2'

    if ':' in str(colNumber):
        se = [int(val) if val else None for val in colNumber.split(':')]
        print(se)
        colNumber = slice(se[0], se[1])
    else:
        try:
            colNumber = int(colNumber)
        except:
            colNumber = 0


def prepareForchekingInDir(dirPath):
    for file in tqdm(os.listdir(dirPath), desc='Обработка файлов'):
        filePath = os.path.join(dirPath, file)
        wb = openpyxl.load_workbook(filePath)
        sh = wb.worksheets[0]
        for row in sh.iter_rows(min_row=3):
            if not row[0]:
                print(f'Есть пустые гуиды - {file}')
                continue
        sh.insert_cols(1)
        sh.insert_cols(1)
        wb.save(filePath)


def writeResultsInXL(results, title='столбец', pathFile: str = "C:\\Users\zaitsev_ad\Desktop\Сохраненные сведения.xlsx",
                     sheetIndex: int = 0, sheetTitle: str = 'Sheet', numerateResults: bool = False):
    if title is None:
        title = ['столбец1', 'столбец2']
    if os.path.exists(pathFile):
        wb = openpyxl.load_workbook(pathFile)
        if len(wb.worksheets) < sheetIndex + 1:
            wb.create_sheet(
                title=sheetTitle,
                # index=sheetIndex
            )
        sh = wb.worksheets[sheetIndex]
        sh.title = sheetTitle
    else:
        wb = openpyxl.Workbook()
        if len(wb.worksheets) < sheetIndex + 1:
            wb.create_sheet(
                title=sheetTitle,
                # index=sheetIndex
            )
        sh = wb.worksheets[sheetIndex]
        sh.title = sheetTitle
    if isinstance(title, list) or isinstance(title, tuple):
        # print(result)
        sh.append([r if not isinstance(r, list) else r[0] for r in title])
    else:
        sh.append([title])

    for n, result in enumerate(results):
        # print(f'{result} --- {type(result)}')
        if isinstance(result, list) or isinstance(result, tuple):
            # print(result)
            sh.append([[n + 1, *r] if numerateResults else r if not isinstance(r, list) else r[0] for r in result])
        else:
            sh.append([n + 1, result] if numerateResults else [result])
    wb.save(pathFile)


def merge_tu(given_tu_list):
    from Dictionary import tuRegExp, tuList
    import re
    print(len(given_tu_list))

    notExists = []
    exists = []
    countExists = {}
    for tu in given_tu_list:
        for reg in tuRegExp:
            # if tu == 'Смоленская обл для ЕРВК26_09_23 (1)':
            #     print(reg, re.search(reg, str(tu).lower().strip()))
            if re.search(reg, str(tu).lower().strip()):
                exists.append(tuRegExp[reg])
                break

    for t in tuList:
        if t not in exists:
            notExists.append(t)
    for e in exists:
        countExists[e] = exists.count(e)
    return {'notExists': notExists, 'exists': countExists, 'countNotExists': len(notExists),
            'countExists': len(countExists)}


if __name__ == '__main__':
    functions = {

        'dellnull': {'action': dellnull,
                     'desc': 'удаляет из указанной папки все файлы, где во второй А2 стоит "Идентификатор ЕИАС", а в А3 пусто, принимает путь до папки',
                     'args': ["Путь до папки"]},
        'bringCol': {'action': bringCol,
                     'desc': 'возвращает список значений указанного столбца по параметрам. Принимает путь до файла',
                     'args': ['Путь до файла', "номер столбца, с которого работаем", "номер строки, с которой работаем",
                              "номер столбца или срез, которые возвращаются"]},
        'changeURaddress': {'action': changeURaddressXL,
                            'desc': 'удаляет из указанной папки все файлы, где во второй А2 стоит "Идентификатор ЕИАС", а в А3 пусто. Принимает путь до файла',
                            'args': ["Путь до файла"]},
        'prepare_for_check': {'action': prepareForchekingInDir,
                              'desc': 'Готовит файл к проверке. Проверяет наличие пустых А1 и добавляет 2 столбца спереди. Принимает путь до файла',
                              'args': ["Путь до файла"]},

    }

    arg1 = argv[1]
    if arg1 in functions:
        # print([None if len(argv) < n+1 else 1 for n in range(functions[arg1]['args'])])
        args = [argv[n + 2] if len(argv) > n + 2 else None for n in range(len(functions[arg1]['args']))]
        if args:
            pprint(functions[arg1]['action'](*args))
        else:
            pprint(functions[arg1]['action']())

    if arg1 == 'help':
        print('------------------------------')
        for name, info in functions.items():
            lenName = len(name)
            middleTire = f"{' ' * (lenName + 2)}{'-' * (30 - lenName)}"
            print(f'{name}  |  {info["desc"]}')
            print(middleTire)
            print(f"{' ' * (lenName + 2)}Аргументы:")
            print(middleTire)

            for n, _arg in enumerate(info["args"]):
                print(f"{' ' * (lenName + 5)}{n + 1}.) {_arg}")
                print(middleTire)
            print('------------------------------')
