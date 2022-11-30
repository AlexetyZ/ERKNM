from ERKNM import erknm
import datetime
import logging
import openpyxl


logging.basicConfig(format='%(asctime)s - [%(levelname)s] - %(name)s - %(funcName)s(%(lineno)d) - %(message)s', filename=f'logging/{datetime.date.today().strftime("%d.%m.%Y")}.log', encoding='utf-8', level=logging.DEBUG)
logger = logging.getLogger(__name__)


def exel_to_json_Objects():
    wb_path = "Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx"
    wb = openpyxl.load_workbook(wb_path)

    # юридические лица
    ul = wb.worksheets[0]

    # отражение нормативной документации
    doc_proverki_ND = wb.worksheets[1]
    viesd_proverki_ND = wb.worksheets[2]
    III_proverki_ND = wb.worksheets[3]

    lenBB = len(ul['B:B'])

    UL_with_objects = {}
    for i in range(4, lenBB + 1):
        if ul[f'S{i}'].value != None:
            continue

        objects = {}

        if ul[f'D{i}'].value != None:
            subject = ul[f'A{i}'].value + '|' + ul[f'D{i}'].value
            objects.update({ul[f'E{i}'].value: {
                'адрес': str.strip(str(ul[f'G{i}'].value)),
                'начало проверки': str(ul[f'B{i}'].value),
                'окончание проверки': str(ul[f'C{i}'].value),
                "ОГРН": str(ul[f'L{i}'].value),
                "деятельность": str(ul[f'N{i}'].value),
                'категория риска': str(ul[f'P{i}'].value),
                'класс опасности': str(ul[f'Q{i}'].value),
                'дата последнего планового кнм': str(ul[f'R{i}'].value),
                'строчка': i}})
            UL_with_objects[subject] = objects

        if ul[f'D{i}'].value == None:
            # print(len(UL_with_objects))
            # print(list(UL_with_objects.values())[len(UL_with_objects)-1])



            # objects.update(list(UL_with_objects.values())[len(UL_with_objects) - 1])



            try:
                objects.update(list(UL_with_objects.values())[len(UL_with_objects) - 1])
            except:
                ul.cell(row=i, column=19, value=ul[f'S{i-1}'].value)
                ul.cell(row=i, column=20, value=ul[f'T{i-1}'].value)
                wb.save(wb_path)
                try:
                    objects.update(list(UL_with_objects.values())[len(UL_with_objects) - 1])
                except Exception as ex:
                    print(ex)
            object_type = ul[f'E{i}'].value
            while True:
                if object_type in objects:
                    object_type = object_type+'1'
                    continue
                if object_type not in objects:
                    break

            #
            objects.update({object_type: {
                'адрес': str.strip(str(ul[f'G{i}'].value)),
                'начало проверки': str(ul[f'B{i}'].value),
                'окончание проверки': str(ul[f'C{i}'].value),
                "ОГРН": str(ul[f'L{i}'].value),
                "деятельность": str(ul[f'N{i}'].value),
                'категория риска': str(ul[f'P{i}'].value),
                'класс опасности': str(ul[f'Q{i}'].value),
                'дата последнего планового кнм': str(ul[f'R{i}'].value),
                'строчка': i}})
            UL_with_objects[list(UL_with_objects.keys())[len(UL_with_objects) - 1]] = objects
            pass
    return UL_with_objects


def main():
    logger.info('Начало выполнения программы')
    try:
        knms = exel_to_json_Objects()
    except:
        knms = exel_to_json_Objects()


    print(knms)




    erknm().enter_knms_in_erknm(knms)




    logger.info('Окончание выполнения программы')



if __name__ == '__main__':
    while True:
        try:
            main()
            break
        except Exception as ex:
            logger.warning(f'ОШИБКА ИСПОЛНЕНИЯ ПРОГРАММЫ! {ex}')
            continue


