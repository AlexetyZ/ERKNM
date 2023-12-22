import openpyxl
from merge_tu import getActualName
from pprint import pprint


def getReasonsDict(path):
    wb = openpyxl.load_workbook(path)
    sh = wb.worksheets[0]
    result = {}
    for row in sh.iter_rows(min_row=1, min_col=1, values_only=True):
        # print(row)
        result[getActualName(row[0])] = dict(sorted(((sh.cell(row=1, column=n + 3).value, cell) for n, cell in enumerate(row[1:])), key=lambda x:x[1], reverse=True))
    return result


def addictionToXl(path, tuCurrentCellNumber, _dict, passCol=1):
    wb = openpyxl.load_workbook(path)
    sh = wb.worksheets[0]
    for row in sh.iter_rows(min_row=2, max_col=sh.max_column+len(_dict)+1):
        tuVal = row[tuCurrentCellNumber].value
        if tuVal in _dict:
            vals = list(_dict[tuVal].items())[passCol-1:passCol][0]
            row[tuCurrentCellNumber+passCol].value = f'{vals[1]}'
    wb.save(path)


if __name__ == '__main__':
    pathReasons = "C:\\Users\zaitsev_ad\Desktop\Сохраненные сведения.xlsx"
    reasonsDict = getReasonsDict(pathReasons)
    print(reasonsDict)
    # for k, v in reasonsDict.items():
    #     vals = list(v.items())[:1][0]
    #     print(k, vals[0], vals[1])
    pathF = "S:\Зайцев_АД\План 2024\этап планирования\ежедневный мониторинг процесса согласования\Перерасчет только деятельность баров и пр.xlsx"
    addictionToXl(
        path=pathF,
        tuCurrentCellNumber=17,
        passCol=1,
        _dict=reasonsDict,
    )
