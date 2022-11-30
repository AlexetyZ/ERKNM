import os.path
import datetime
import re
import fake_useragent
import undetected_chromedriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
import time

import Proverochnii_list
from config import erknm_accounts
import logging
from functools import partial
import Normative_documentation
import openpyxl
import clipboard

logging.basicConfig(format='%(asctime)s - [%(levelname)s] - %(name)s - %(funcName)s(%(lineno)d) - %(message)s',
                    filename=f'logging/{datetime.date.today().strftime("%d.%m.%Y")}.log', encoding='utf-8',
                    level=logging.INFO)
logger = logging.getLogger(__name__)


class erknm:
    def __init__(self, headless: bool = False):
        self.options = undetected_chromedriver.ChromeOptions()
        prefs = {'profile.default_content_setting_values': {'images': 2,
                                                            'plugins': 2, 'popups': 2, 'geolocation': 2,
                                                            'notifications': 2, 'auto_select_certificate': 2,
                                                            'mouselock': 2, 'mixed_script': 2, 'media_stream': 2,
                                                            'media_stream_mic': 2, 'media_stream_camera': 2,
                                                            'protocol_handlers': 2,
                                                            'ppapi_broker': 2, 'automatic_downloads': 2,
                                                            'midi_sysex': 2,
                                                            'push_messaging': 2, 'ssl_cert_decisions': 2,
                                                            'metro_switch_to_desktop': 2,
                                                            'protected_media_identifier': 2, 'app_banner': 2,
                                                            'site_engagement': 2,
                                                            'durable_storage': 2}}
        self.options.add_experimental_option('prefs', prefs)
        self.options.add_argument("--disable-native-events")
        self.options.add_argument("disable-infobars")
        self.options.add_argument("--disable-extensions")
        if headless is True:
            self.options.add_argument("--headless")  # если это включено, то будет в режиме фантома

        # host_dir = 'C:\\Users\\user\\AppData\\Local\\Google\\Chrome\\User Data\\Default'
        # self.options.add_argument('user-data-dir=' + host_dir)
        self.browser = undetected_chromedriver.Chrome(self.options)

        logger.info('Начало авторизации')
        # self.Controller(partial(self.autorize))

    def Controller(self, function, extraArgs=''):  # 'Эта функция гарантирует выполнение процессов, давая им 5 попыток.

        popit = 5
        # autorization = 0
        while popit != 0:
            try:

                res = function(*extraArgs)
                if re.search('add_knm', str(function)):
                    return res
                logger.info(f'{str(function)} прошла с {6 - popit} раза')

                # autorization = 1
                break
            except Exception as ex:
                logger.warning(f"Ошибка авторизации: {function} {ex}")

                popit -= 1
                if popit == 0:
                    logger.info(
                        f'{str(function)} не удалась, имеется непредвиденная ошибка, было предпринято 5 попыток')

                    self.browser.quit()
                    return f'{str(function)} не удалась, имеется непредвиденная ошибка, было предпринято 5 попыток'

    def autorize(self):
        self.browser.get('https://private.proverki.gov.ru/private/auth')

        try:
            WebDriverWait(self.browser, 10).until(EC.presence_of_element_located(
                (By.XPATH, '//*[@id="root"]/div/main/div/form/div[2]/button[2]'))).click()
        except Exception as ex:
            logger.warning(ex)
            try:
                self.browser.find_element(by=By.XPATH, value='//*[@id="details-button"]').click()
                time.sleep(1)
                self.browser.find_element(by=By.XPATH, value='//*[@id="proceed-link"]').click()
                time.sleep(2)
                WebDriverWait(self.browser, 10).until(EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="root"]/div/main/div/form/div[2]/button[2]'))).click()

            except Exception as ex:
                pass
            logger.warning(f'{ex=}')



        WebDriverWait(self.browser, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="login"]'))).send_keys(
            str(erknm_accounts['Alexety']['login']))
        self.browser.find_element(by=By.XPATH, value='//*[@id="password"]').send_keys(
            str(erknm_accounts['Alexety']['password']))
        self.browser.find_element(by=By.XPATH, value='/html/body/esia-root/div/esia-login/div/div[1]/form/div[3]/button').click()
        WebDriverWait(self.browser, 10).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div/div/header/div/div[2]/button[1]')))
        ekrnm_button_class = self.browser.find_element(by=By.XPATH,
                                                       value='/html/body/div/div/header/div/div[2]/button[1]')

        while ekrnm_button_class.get_attribute(
                'class') == '_Button_5bv2s_1 _ButtonDefault_5bv2s_55 _SwitchButton_8ofrs_1 ButtonDefault _ButtonLarge_5bv2s_31 _ButtonContained_5bv2s_15':
            ekrnm_button_class.click()
        self.browser.get('https://private.proverki.gov.ru/private/knms')
        logger.info('Autorization successfull completed!')
        # time.sleep(50000)

    #


    def enter_knms_in_erknm(self, knms):
        logger.info('Ready knms list to enter in erknm')

        for knm in knms.items():
            if f'{knm[0].split("|")[0]}' == 'рейдовая':
                result = self.add_knm_reid(knm)

            if f'{knm[0].split("|")[0]}' == 'выездная':
                result = self.add_knm_visited(knm)

            if f'{knm[0].split("|")[0]}' == 'документарная':
                result = self.add_knm_documentary(knm)

            logger.info(f'{knm} has been entered in erknm!')

            subject_data = list(knm[1].items())
            print(subject_data[0][1]['строчка'])
            wb = openpyxl.load_workbook("Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx")
            # юридические лица
            row = int(subject_data[0][1]['строчка'])
            ul = wb.worksheets[0]
            ul.cell(row=row, column=19, value=result['number'])
            ul.cell(row=row, column=20, value=result['status'])
            logger.info(f"номер кнм {knm[0]}  --->>>  {result['number']}, status   --->>>  {result['status']}")

            wb.save("Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx")

        logger.info('All knms has been entered in erknm!')

    def add_knm_documentary(self, knm):

        self.browser.get('https://private.proverki.gov.ru/private/knm/new')
        subject_data = list(knm[1].items())
        len_objects = len(subject_data)

        logger.info('')
        logger.info({knm[0]})
        logger.info('')

        object_kinds = {
            'дошкольные образовательные организации': '174',
            'общеобразовательные организации': '175',
            'Деятельность детских лагерей на время каникул': '252',
            'школы-интернаты, специальные (коррекционные) общеобразовательные организации': '182',
            'перинатальные центры, родильные дома, родильные отделения': '162',
            'Объекты здравоохранения': '0',
            'дома (интернаты) для лиц с физическими или умственными недостатками, в том числе геронтопсихиатрические центры, психоневрологические интернаты': '161'

        }

        objects_risk = {
            'чрезвычайно высокий риск': '0',
            'высокий риск': '1',
            'значительный риск': '2',
            'средний риск': '3',
            'умеренный риск': '4',
        }

        object_danger = {
            'Первый': '0',
            'Второй': '1',
            'Третий': '2',
            'Четвертый': '3',
        }

        # принцип формирования объектов выбора: при открытии параметра выбора, добавляется див в конце страницы, чтобы узнать его айди, используй конструкцию print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML')), чтобы сделать снимок страницы в момент выпадания списка, ибо иначе его никак не захватишь, потом ища ключевую фразу( она будет последней), и смотри айдишник, на который будешь кликать

        # Наименование органа контроля

        logger.info('вставляем наименование органа контроля')
        WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div/div/main/form/div[2]/section[1]/div[2]/div/div[2]/div[1]/div/div[1]')))
        time.sleep(1)
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[2]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value='react-select-2-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка наименования органа контроля не прошла: {ex}')
        logger.info('вставка наименования органа контроля прошла успешно')

        # Вид контроля

        logger.info('вставляем Вид контроля')
        WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div/div/main/form/div[2]/section[1]/div[3]/div/div[2]/div[1]/div/div[1]')))
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[3]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value='react-select-3-option-2').click()
        except Exception as ex:
            logger.warning(f'вставка вид контроля не прошла: {ex}')
        logger.info('вставка Вид контроля прошла успешно')

        # Вид КНМ

        logger.info('вставляем Вид КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[1]/div[4]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[4]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value=f'react-select-4-option-5').click()
        except Exception as ex:
            logger.warning(f'вставка Вид КНМ не прошла: {ex}')
        logger.info('вставка Вид КНМ прошла успешно')

        # Характер КНМ
        self.enter_type_proverka(5)

        # Номер плана
        # time.sleep(100000)

        logger.info('вставляем Номер плана')
        self.wait_loader(WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.ID, 'planId'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.ID,
                                      value='planId').send_keys('2023050341')
        except Exception as ex:
            logger.warning(f'вставка Номер плана не прошла: {ex}')
        logger.info('вставка Номер плана прошла успешно')

        # Дата начала КНМ

        logger.info('вставляем Дата начала КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="startDateBlock"]/div[2]/div[1]/div/div/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="startDateBlock"]/div[2]/div[1]/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['начало проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                    '%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата начала КНМ не прошла: {ex}')
        logger.info('вставка Дата начала КНМ прошла успешно')

        # Cрок проведения (дней)

        logger.info('вставляем Cрок проведения (дней)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="durationDaysBlock"]/div[2]/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="durationDaysBlock"]/div[2]/input').send_keys('10')
        except Exception as ex:
            logger.warning(f'вставка Cрок проведения (дней) не прошла: {ex}')
        logger.info('вставка Cрок проведения (дней) прошла успешно')

        # Срок непосредственного взаимодействия (часов)

        logger.info('вставляем Срок непосредственного взаимодействия (часов)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="durationHoursBlock"]/div[2]/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="durationHoursBlock"]/div[2]/input').send_keys('80')
        except Exception as ex:
            logger.warning(f'вставка Срок непосредственного взаимодействия (часов) не прошла: {ex}')
        logger.info('вставка Срок непосредственного взаимодействия (часов) прошла успешно')

        # Основания включения в план

        logger.info('вставляем Основания включения в план')
        # жмем кнопку добавить
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.ID, "addReasonButton"))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.ID, value="addReasonButton").click()

        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div[1]/div/main/form/div[2]/section[1]/div[13]/div[2]/div/div/div/div[1]/div/div[1]/div[1]'))))  # используем wait_loader, так как страница запускает лоадер

        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[13]/div[2]/div/div/div/div[1]/div/div[1]/div[1]').click()

        try:
            self.browser.find_element(by=By.ID, value='react-select-9-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка Основания включения в план не прошла: {ex}')
            # Не забыть вставить дату окончания последнего планового КНМ

        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="reasonsBlock"]/div[2]/div/div/div[2]/div/div[1]/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['дата последнего планового кнм'],
                                           '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y'))
            logger.info('вставка Основания включения в план прошла успешно')
        except Exception as ex:
            logger.warning(f'вставка Дата дата последнего планового КНМ не прошла: {ex}')

        # Наименование прокуратуры

        logger.info('вставляем Наименование прокуратуры')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[2]/div[2]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.wait_loader(self.browser.find_element(by=By.XPATH,
                                                   value='/html/body/div/div/main/form/div[2]/section[2]/div[2]/div/div[2]/div[1]/div/div[1]').click())
        try:
            self.browser.find_element(by=By.ID, value='react-select-7-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка Наименование прокуратуры не прошла: {ex}')
        logger.info('вставка Наименование прокуратуры прошла успешно')

        # Основной государственный регистрационный номер (ОГРН)

        logger.info('вставляем Основной государственный регистрационный номер (ОГРН)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="organizations[0].ogrn"]'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="organizations[0].ogrn"]').send_keys(subject_data[0][1]['ОГРН'])
        except Exception as ex:
            logger.warning(f'вставка Основной государственный регистрационный номер (ОГРН) не прошла: {ex}')

        logger.info('вставка Основной государственный регистрационный номер (ОГРН) прошла успешно')

        # ждем, пока предложит варианты выпадающего списка
        try:
            WebDriverWait(self.browser, 30).until(EC.presence_of_element_located((By.ID, 'autoCompleteList')))
            self.browser.find_element(by=By.ID, value='autoCompleteList').click()
        except Exception as ex:
            logger.warning(f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) не прошел: {ex}')

        logger.info(f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) прошел успешно!')

        # Добавляем количество объектов

        for sd in range(len_objects - 1):
            self.browser.find_element(by=By.ID, value='erknmObjectsAddButton').click()
            time.sleep(0.25)

        for n, subj_dat in enumerate(subject_data):

            # ставим местонахождение при необъодимости
            if n != 0:
                clipboard.copy(subj_dat[1]['адрес'])
                self.browser.find_element(by=By.XPATH,
                                          value=f'/html/body/div[1]/div/main/form/div[2]/section[3]/div[3]/div[2]/div[{n + 1}]/div/div[2]/div[1]/div/div[2]/div/textarea').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')

            # выбираем тип для объекта
            try:
                logger.info(f'выбираем тип для объекта {subj_dat[n]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].objectType').click()
                self.browser.find_element(by=By.ID, value=f'react-select-{10 + (n * 5)}-option-0').click()
                logger.info(f'выбор тип для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор тип для объекта {subj_dat[0]} не прошел: {ex}')

            time.sleep(0.25)

            # выбираем вид для объекта
            # print(f'{object_kinds[subj_dat[1]["деятельность"]]}=')
            try:
                logger.info(f'выбираем вид для объекта {subj_dat[n]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].objectKind').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{11 + (n * 5)}-option-{object_kinds[subj_dat[1]["деятельность"]]}').click()
                logger.info(f'выбор вид для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор вид для объекта {subj_dat[0]} не прошел: {ex}')
            time.sleep(0.25)
            # print('закончил вид объекта')

            # выбираем Категория риска для объекта
            # print(f'{objects_risk[subj_dat[1]["категория риска"]]=}')
            try:
                logger.info(f'выбираем Категория риска для объекта {subj_dat[n]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].riskCategory').click()

                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{13 + (n * 5)}-option-{objects_risk[subj_dat[1]["категория риска"]]}').click()
                logger.info(f'выбор Категория риска для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор Категория риска для объекта {subj_dat[0]} не прошел: {ex}')
            time.sleep(0.25)

            # выбираем Класс опасности для объекта
            try:
                logger.info(f'выбираем Класс опасности для объекта {subj_dat[0]}')

                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].dangerClass').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{14 + (n * 5)}-option-{object_danger[subj_dat[1]["класс опасности"]]}').click()
                logger.info(f'выбор Класс опасности для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор Класс опасности для объекта {subj_dat[0]} не прошел: {ex}')
            time.sleep(0.25)

        # добавляем Перечень действий, осуществляемый в рамках КНМ

        examination_ivents = ['0', '1', ]  # айдишники типов действий, осуществляемых в рамках КНМ

        try:
            for l in examination_ivents:
                self.browser.find_element(by=By.ID, value='erknmEventsAddButton').click()
                time.sleep(0.2)
            for n, lei in enumerate(examination_ivents):
                self.browser.find_element(by=By.ID, value=f'eventsErknm[{n}].type').click()

                # print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))
                # time.sleep(10000)

                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{10 + (len_objects * 5) + n}-option-{lei}').click()
                self.browser.find_element(by=By.XPATH,
                                          value=f'//*[@id="eventsBlock"]/div[2]/div[{n + 1}]/div[2]/div/div[1]/div/div/input').send_keys(
                    datetime.datetime.strptime(subject_data[0][1]['начало проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%d.%m.%Y'))
                self.browser.find_element(by=By.XPATH,
                                          value=f'//*[@id="eventsBlock"]/div[2]/div[{n + 1}]/div[3]/div/div[1]/div/div/input').send_keys(
                    datetime.datetime.strptime(subject_data[0][1]['окончание проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%d.%m.%Y'))

            logger.info(f'Добавлено Перечень действий, осуществляемый в рамках КНМ прошел успешно')
        except Exception as ex:
            logger.warning(f'Добавление Перечень действий, осуществляемый в рамках КНМ не прошло: {ex}')

        # print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))

        # заполняем Обязательные требования, подлежащие проверке

        ND = getattr(Normative_documentation, f'_{object_kinds[subject_data[0][1]["деятельность"]]}')
        for nd in ND:
            self.browser.find_element(by=By.ID, value='erknmRequirementsAddButton').click()
            time.sleep(0.1)
        logger.info(f'Добавлено Перечень действий, осуществляемый в рамках КНМ прошел успешно')
        requirements = self.browser.find_elements(by=By.XPATH,
                                                  value='/html/body/div[1]/div/main/form/div[2]/section[4]/div[2]/table/tbody/tr')

        for n, req in enumerate(requirements):
            try:
                req.find_element(by=By.XPATH, value='./td[2]/div/div').click()

                time.sleep(1)
                formular_npa = ND[n]['ФОРМУЛИРОВКА ТРЕБОВАНИЯ']
                name_npa = ND[n]['НАИМЕНОВАНИЕ НПА']

                req.find_element(by=By.XPATH,
                                 value=f'./td[2]/div/div/div[2]//div[contains(text()="{formular_npa}")]').click()
                time.sleep(1)
                print('короткий путь')


            except Exception as ex:

                req.find_element(by=By.XPATH, value='./td[2]/div/div/div[2]//div[text()="Создать новый"]').click()
                time.sleep(1)
                clipboard.copy(ND[n]['ФОРМУЛИРОВКА ТРЕБОВАНИЯ'])

                req.find_element(by=By.XPATH, value='./td[2]/div[2]//textarea').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
                time.sleep(2)
                clipboard.copy(ND[n]['НАИМЕНОВАНИЕ НПА'])
                req.find_element(by=By.XPATH, value='./td[3]/div[1]//textarea').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
                print('длинный путь')
                time.sleep(2)
                clipboard.copy(ND[n]['ДАТА НПА'])
                req.find_element(by=By.XPATH, value='./td[4]/div[1]//input').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
                time.sleep(1)

        # Добавляем места проведения контрольного (надзорного) мероприятия

        logger.info('Добавляем места проведения контрольного (надзорного) мероприятия')
        try:
            self.browser.find_element(by=By.ID, value='erknmPlacesAddButton').click()
            logger.info('Добавление места проведения контрольного (надзорного) мероприятия прошла успешно')

        except Exception as ex:
            logger.warning(f'Добавление места проведения контрольного (надзорного) мероприятия не прошла: {ex}')

        time.sleep(0.5)

        # Вставляем места проведения контрольного (надзорного) мероприятия

        logger.info('Вставляем места проведения контрольного (надзорного) мероприятия')
        try:
            self.browser.find_element(by=By.XPATH,
                                      value=f'/html/body/div[1]/div/main/form/div[2]/section[4]/div[5]/div[2]/div/div/div/textarea').send_keys(
                'г.Нижний Новгород, пр.Ильича, д.3')

        except Exception as ex:
            logger.warning(f'Вставляем места проведения контрольного (надзорного) мероприятия не прошла: {ex}')
        logger.info('Вставляем места проведения контрольного (надзорного) мероприятия прошла успешно')
        time.sleep(1)

        # сохраняем проверку

        try:
            self.wait_loader(self.browser.find_element(by=By.ID, value='saveButton').click())

        except Exception as ex:
            logger.warning(f'Сохранение проверки не прошло: {ex}')
        logger.info('Сохранение проверки прошло успешно')
        time.sleep(0.5)

        print('Завершено')
        time.sleep(5)
        # Получение номера и статуса проверки кнм
        WebDriverWait(self.browser, 100).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/header/div/div[1]/div')))
        try:
            self.wait_loader(self.browser.find_element(by=By.TAG_NAME, value='title'))
            time.sleep(5)
            try:
                knm_number = self.browser.find_element(by=By.CLASS_NAME, value='KnmInfo_Title__LQcuP').text
            except:
                knm_number = self.browser.find_element(by=By.CLASS_NAME, value='KnmHeader_Title__vO5Qr').text
            try:
                knm_status = self.browser.find_element(by=By.CLASS_NAME, value='KnmInfo_Status__3ZqpQ').text
            except:
                try:
                    knm_status = self.browser.find_element(by=By.CLASS_NAME, value='KnmHeader_Status__118Oo').text
                except:
                    knm_status = 'не удалось подгрузить статус'

            result = {'number': knm_number, 'status': knm_status}
            print(result)
            logger.info('Получение номера и статуса проверки кнм')

            return result
        except Exception as ex:
            logger.warning(f'Получение номера и статуса проверки кнм не прошла: {ex}')

            result = {'number': None, 'status': None}

            print(result)
            return result

        # print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))

    def add_knm_reid(self, knm):

        self.browser.get('https://private.proverki.gov.ru/private/knm/new')
        subject_data = list(knm[1].items())
        len_objects = len(subject_data)

        logger.info('')
        logger.info(knm[0])
        logger.info('')

        object_kinds = {
            'дошкольные образовательные организации': '174',
            'общеобразовательные организации': '175',
            'Деятельность детских лагерей на время каникул': '252',
            'школы-интернаты, специальные (коррекционные) общеобразовательные организации': '182',
            'перинатальные центры, родильные дома, родильные отделения': '162',
            'Объекты здравоохранения': '0',
            'дома (интернаты) для лиц с физическими или умственными недостатками, в том числе геронтопсихиатрические центры, психоневрологические интернаты': '161'

        }

        objects_risk = {
            'чрезвычайно высокий риск': '0',
            'высокий риск': '1',
            'значительный риск': '2',
            'средний риск': '3',
            'умеренный риск': '4',
        }

        object_danger = {
            'Первый': '0',
            'Второй': '1',
            'Третий': '2',
            'Четвертый': '3',
        }

        # принцип формирования объектов выбора: при открытии параметра выбора, добавляется див в конце страницы, чтобы узнать его айди, используй конструкцию print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML')), чтобы сделать снимок страницы в момент выпадания списка, ибо иначе его никак не захватишь, потом ища ключевую фразу( она будет последней), и смотри айдишник, на который будешь кликать

        # Наименование органа контроля

        logger.info('вставляем наименование органа контроля')
        WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div/div/main/form/div[2]/section[1]/div[2]/div/div[2]/div[1]/div/div[1]')))
        time.sleep(1)
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[2]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value='react-select-2-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка наименования органа контроля не прошла: {ex}')
        logger.info('вставка наименования органа контроля прошла успешно')

        # Вид контроля

        logger.info('вставляем Вид контроля')
        WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div/div/main/form/div[2]/section[1]/div[3]/div/div[2]/div[1]/div/div[1]')))
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[3]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value='react-select-3-option-2').click()
        except Exception as ex:
            logger.warning(f'вставка вид контроля не прошла: {ex}')
        logger.info('вставка Вид контроля прошла успешно')

        # Вид КНМ

        logger.info('вставляем Вид КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[1]/div[4]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[4]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value=f'react-select-4-option-4').click()
        except Exception as ex:
            logger.warning(f'вставка Вид КНМ не прошла: {ex}')
        logger.info('вставка Вид КНМ прошла успешно')

        # Характер КНМ

        self.enter_type_proverka(5)

        # Номер плана
        # time.sleep(100000)

        logger.info('вставляем Номер плана')
        self.wait_loader(WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.ID, 'planId'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.ID,
                                      value='planId').send_keys('2023050341')
        except Exception as ex:
            logger.warning(f'вставка Номер плана не прошла: {ex}')
        logger.info('вставка Номер плана прошла успешно')

        # Дата начала КНМ

        logger.info('вставляем Дата начала КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="startDateBlock"]/div[2]/div[1]/div/div/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="startDateBlock"]/div[2]/div[1]/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['начало проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                    '%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата начала КНМ не прошла: {ex}')
        logger.info('вставка Дата начала КНМ прошла успешно')

        # Дата окончания КНМ

        logger.info('вставляем Дата окончания КНМ')
        self.wait_loader(WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="stopDateBlock"]/div[2]/div/div/div/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="stopDateBlock"]/div[2]/div/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['окончание проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                    '%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата окончания КНМ не прошла: {ex}')
        logger.info('вставка Дата окончания КНМ прошла успешно')

        # Cрок проведения (дней)

        logger.info('вставляем Cрок проведения (дней)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="durationDaysBlock"]/div[2]/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="durationDaysBlock"]/div[2]/input').send_keys('10')
        except Exception as ex:
            logger.warning(f'вставка Cрок проведения (дней) не прошла: {ex}')
        logger.info('вставка Cрок проведения (дней) прошла успешно')

        # Срок непосредственного взаимодействия (часов)

        logger.info('вставляем Срок непосредственного взаимодействия (часов)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="durationHoursBlock"]/div[2]/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="durationHoursBlock"]/div[2]/input').send_keys('80')
        except Exception as ex:
            logger.warning(f'вставка Срок непосредственного взаимодействия (часов) не прошла: {ex}')
        logger.info('вставка Срок непосредственного взаимодействия (часов) прошла успешно')

        # Основания включения в план

        logger.info('вставляем Основания включения в план')
        # жмем кнопку добавить
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.ID, "addReasonButton"))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.ID, value="addReasonButton").click()

        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[1]/div[13]/div[2]/div/div/div/div[1]/div/div[1]/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[13]/div[2]/div/div/div/div[1]/div/div[1]/div[1]').click()

        try:
            self.browser.find_element(by=By.ID, value='react-select-9-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка Основания включения в план не прошла: {ex}')
            # Не забыть вставить дату окончания последнего планового КНМ

        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="reasonsBlock"]/div[2]/div/div/div[3]/div/div[1]/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['дата последнего планового кнм'],
                                           '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата дата последнего планового КНМ не прошла: {ex}')

        logger.info('вставка Основания включения в план прошла успешно')

        # Наименование прокуратуры

        logger.info('вставляем Наименование прокуратуры')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[2]/div[2]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.wait_loader(self.browser.find_element(by=By.XPATH,
                                                   value='/html/body/div/div/main/form/div[2]/section[2]/div[2]/div/div[2]/div[1]/div/div[1]').click())
        try:
            self.browser.find_element(by=By.ID, value='react-select-7-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка Наименование прокуратуры не прошла: {ex}')
        logger.info('вставка Наименование прокуратуры прошла успешно')

        # Основной государственный регистрационный номер (ОГРН)

        logger.info('вставляем Основной государственный регистрационный номер (ОГРН)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="organizations[0].ogrn"]'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="organizations[0].ogrn"]').send_keys(subject_data[0][1]['ОГРН'])
        except Exception as ex:
            logger.warning(f'вставка Основной государственный регистрационный номер (ОГРН) не прошла: {ex}')

        logger.info('вставка Основной государственный регистрационный номер (ОГРН) прошла успешно')

        # ждем, пока предложит варианты выпадающего списка
        try:
            WebDriverWait(self.browser, 30).until(EC.presence_of_element_located((By.ID, 'autoCompleteList')))
            self.browser.find_element(by=By.ID, value='autoCompleteList').click()
        except Exception as ex:
            logger.warning(f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) не прошел: {ex}')

        logger.info(f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) прошел успешно!')

        # Добавляем один объект

        self.browser.find_element(by=By.ID, value='erknmObjectsAddButton').click()
        time.sleep(0.25)

        # выбираем тип для объекта
        try:
            logger.info(f'выбираем тип для объекта {subject_data[0][1]}')
            self.browser.find_element(by=By.ID, value=f'objectsErknm[0].objectType').click()
            self.browser.find_element(by=By.ID, value=f'react-select-10-option-0').click()
            logger.info(f'выбор тип для объекта {subject_data[0][1]} прошел успешно')


        except Exception as ex:
            logger.warning(f'выбор тип для объекта {subject_data[0][1]} не прошел: {ex}')

            time.sleep(0.25)

        # выбираем вид для объекта

        kind_object = object_kinds[subject_data[0][1]["деятельность"]]
        print(kind_object)
        try:
            logger.info(f'выбираем вид для объекта {subject_data[0][1]}')
            self.browser.find_element(by=By.ID, value=f'objectsErknm[0].objectKind').click()
            self.browser.find_element(by=By.ID,
                                      value=f'react-select-11-option-{object_kinds[subject_data[0][1]["деятельность"]]}').click()
            logger.info(f'выбор вид для объекта {subject_data[0][1]} прошел успешно')
        except Exception as ex:
            logger.warning(f'выбор вид для объекта {subject_data[0][1]} не прошел: {ex}')
        time.sleep(0.25)

        # выбираем Категория риска для объекта
        try:
            logger.info(f'выбираем Категория риска для объекта {subject_data[0][1]}')
            self.browser.find_element(by=By.ID, value=f'objectsErknm[0].riskCategory').click()
            # print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))
            self.browser.find_element(by=By.ID,
                                      value=f'react-select-13-option-{objects_risk[subject_data[0][1]["категория риска"]]}').click()
            logger.info(f'выбор Категория риска для объекта {subject_data[0][1]} прошел успешно')
        except Exception as ex:
            logger.warning(f'выбор Категория риска  не прошел для объекта {subject_data[0][1]}: {ex}')
        time.sleep(0.25)

        # выбираем Класс опасности для объекта
        try:
            logger.info(f'выбираем Класс опасности для объекта {subject_data[0][1]}')
            self.browser.find_element(by=By.ID, value=f'objectsErknm[0].dangerClass').click()
            self.browser.find_element(by=By.ID,
                                      value=f'react-select-14-option-{object_danger[subject_data[0][1]["класс опасности"]]}').click()
            logger.info(f'выбор Класс опасности для объекта {subject_data[0][1]} прошел успешно')
        except Exception as ex:
            logger.warning(f'выбор Класс опасности для объекта {subject_data[0][1]} не прошел: {ex}')
        time.sleep(0.25)

        # добавляем Перечень действий, осуществляемый в рамках КНМ

        examination_ivents = ['0', '3', '4', '6', '5', '8', ]  # айдишники типов действий, осуществляемых в рамках КНМ

        try:
            for l in examination_ivents:
                self.browser.find_element(by=By.ID, value='erknmEventsAddButton').click()
                time.sleep(0.2)
            for n, lei in enumerate(examination_ivents):
                self.browser.find_element(by=By.ID, value=f'eventsErknm[{n}].type').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{15 + n}-option-{lei}').click()
                self.browser.find_element(by=By.XPATH,
                                          value=f'//*[@id="eventsBlock"]/div[2]/div[{n + 1}]/div[2]/div/div[1]/div/div/input').send_keys(
                    datetime.datetime.strptime(subject_data[0][1]['начало проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%d.%m.%Y'))
                self.browser.find_element(by=By.XPATH,
                                          value=f'//*[@id="eventsBlock"]/div[2]/div[{n + 1}]/div[3]/div/div[1]/div/div/input').send_keys(
                    datetime.datetime.strptime(subject_data[0][1]['окончание проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%d.%m.%Y'))

            logger.info(f'Добавлено Перечень действий, осуществляемый в рамках КНМ прошел успешно')
        except Exception as ex:
            logger.warning(f'Добавление Перечень действий, осуществляемый в рамках КНМ не прошло: {ex}')

        time.sleep(3)

        # Добавляем оставшиеся контролируемые лица
        try:
            for n, sd in enumerate(subject_data):
                if n != 0:
                    print('lf,fdbv')
                    while True:
                        try:
                            self.browser.find_element(by=By.XPATH,
                                                      value='/html/body/div[1]/div/main/form/div[2]/section[3]/div[1]/span/button').click()
                            break
                        except:
                            for i in range(5):
                                self.browser.find_element(by=By.TAG_NAME, value='html').send_keys(Keys.UP)
                    time.sleep(3)
                    while True:
                        try:
                            self.browser.find_element(by=By.XPATH, value=f'//*[@id="organizations[{n}].ogrn"]')
                            print('пока не создано')
                            break
                        except:
                            time.sleep(2)

            logger.info(f'Добавление Сведения об остальных контролируемых лицах прошло успешно')
        except Exception as ex:
            logger.warning(f'Добавление Сведения об остальных контролируемых лицах не прошло: {ex}')
        # заносим остальные контролируемые лица

        for n, sd in enumerate(subject_data):
            if n != 0:

                logger.info('вставляем Основной государственный регистрационный номер (ОГРН)')
                # self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
                #     (By.XPATH,
                #      f'//*[@id="organizations[{n}].ogrn"]'))))  # используем wait_loader, так как страница запускает лоадер
                try:
                    self.browser.find_element(by=By.XPATH,
                                              value=f'//*[@id="organizations[{n}].ogrn"]').send_keys(sd[1]['ОГРН'])
                    logger.info('вставка Основной государственный регистрационный номер (ОГРН) прошла успешно')
                except Exception as ex:
                    logger.warning(f'вставка Основной государственный регистрационный номер (ОГРН) не прошла: {ex}')

                # ждем, пока предложит варианты выпадающего списка
                try:
                    WebDriverWait(self.browser, 30).until(EC.presence_of_element_located((By.ID, 'autoCompleteList')))
                    self.browser.find_element(by=By.ID, value='autoCompleteList').click()
                    logger.info(
                        f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) прошел успешно!')
                except Exception as ex:
                    logger.warning(
                        f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) не прошел: {ex}')

        # заполняем Обязательные требования, подлежащие проверке

        ND = getattr(Normative_documentation, f'_{object_kinds[subject_data[0][1]["деятельность"]]}')
        for nd in ND:
            self.browser.find_element(by=By.ID, value='erknmRequirementsAddButton').click()
            time.sleep(0.1)
        logger.info(f'Добавлено Перечень действий, осуществляемый в рамках КНМ прошел успешно')
        requirements = self.browser.find_elements(by=By.XPATH,
                                                  value='/html/body/div[1]/div/main/form/div[2]/section[4]/div[2]/table/tbody/tr')

        for n, req in enumerate(requirements):
            try:
                req.find_element(by=By.XPATH, value='./td[2]/div/div').click()

                time.sleep(1)
                formular_npa = ND[n]['ФОРМУЛИРОВКА ТРЕБОВАНИЯ']
                name_npa = ND[n]['НАИМЕНОВАНИЕ НПА']

                req.find_element(by=By.XPATH,
                                 value=f'./td[2]/div/div/div[2]//div[contains(text()="{formular_npa}")]').click()
                time.sleep(1)
                print('короткий путь')


            except Exception as ex:

                req.find_element(by=By.XPATH, value='./td[2]/div/div/div[2]//div[text()="Создать новый"]').click()
                time.sleep(1)
                clipboard.copy(ND[n]['ФОРМУЛИРОВКА ТРЕБОВАНИЯ'])

                req.find_element(by=By.XPATH, value='./td[2]/div[2]//textarea').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
                time.sleep(2)
                clipboard.copy(ND[n]['НАИМЕНОВАНИЕ НПА'])
                req.find_element(by=By.XPATH, value='./td[3]/div[1]//textarea').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
                print('длинный путь')
                time.sleep(2)
                clipboard.copy(ND[n]['ДАТА НПА'])
                req.find_element(by=By.XPATH, value='./td[4]/div[1]//input').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
                time.sleep(1)

        # отметка об использовании проверочного листа

        logger.info('Нажимаем отметка об использовании проверочного листа')
        while True:
            try:
                self.wait_loader(self.browser.find_element(by=By.ID, value='isChecklistsUsed').click())
            except Exception as ex:
                logger.warning(f'Нажатие отметки об использовании проверочного листа не прошла: {ex}')
            logger.info('Нажатие отметки об использовании проверочного листа прошла успешно')
            try:
                self.browser.find_element(by=By.XPATH, value='//*[@id="checklistsBlock"]/div/button')
                break
            except:
                continue
            time.sleep(1)

        # Добавляем проверочный лист

        PL = getattr(Proverochnii_list, f'_{object_kinds[subject_data[0][1]["деятельность"]]}')
        print(len(PL))
        for every_pl in PL:
            logger.info('Добавляем проверочный лист')
            try:
                self.browser.find_element(by=By.XPATH, value='//*[@id="checklistsBlock"]/div/button').click()

            except Exception as ex:
                logger.warning(f'Добавляем проверочный лист не прошла: {ex}')
            logger.info('Добавляем проверочный лист прошла успешно')
            time.sleep(0.5)

        #  выбираем проверочный лист
        logger.info('выбираем проверочный лист')
        for n, pl in enumerate(PL):
            try:
                self.browser.find_element(By.XPATH,
                                          value=f'/html/body/div[1]/div/main/form/div[2]/section[4]/div[6]/div/ul/li[{n + 1}]/div/div[2]/div[2]/div[2]/div/div/div/div[1]').click()
                self.browser.find_element(by=By.XPATH,
                                          value=f'/html/body/div[1]/div/main/form/div[2]/section[4]/div[6]/div/ul/li[{n + 1}]/div/div[2]/div[2]/div[2]/div/div/div/div[1]//div[text()="{pl}"]').click()
                # logger.info(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))

            except Exception as ex:
                logger.warning(f'Выбор проверочных листов не прошел: {ex}')
                print(ex)
            time.sleep(2)
        logger.info('Выбор всех проверочных листов успешно завершен')

        #
        #
        #
        # Добавляем места проведения контрольного (надзорного) мероприятия

        for o in subject_data:
            logger.info('Добавляем места проведения контрольного (надзорного) мероприятия')
            try:
                self.browser.find_element(by=By.ID, value='erknmPlacesAddButton').click()

            except Exception as ex:
                logger.warning(f'Добавление места проведения контрольного (надзорного) мероприятия не прошла: {ex}')
        logger.info('Добавление места проведения контрольного (надзорного) мероприятия прошла успешно')
        time.sleep(0.5)

        # Вставляем места проведения контрольного (надзорного) мероприятия
        for n, obj in enumerate(subject_data):
            logger.info('Вставляем места проведения контрольного (надзорного) мероприятия')
            try:
                self.browser.find_element(by=By.XPATH,
                                          value=f'/html/body/div/div/main/form/div[2]/section[4]/div[6]/div[2]/div[{n + 1}]/div/div/textarea').send_keys(
                    str(obj[1]['адрес']))


            except Exception as ex:
                logger.warning(f'Вставляем места проведения контрольного (надзорного) мероприятия не прошла: {ex}')
        logger.info('Вставляем места проведения контрольного (надзорного) мероприятия прошла успешно')
        time.sleep(0.5)
        #
        #
        #
        # # сохраняем проверку
        #
        #
        # try:
        #     self.browser.find_element(by=By.ID, value='saveButton').click()
        #
        # except Exception as ex:
        #     logger.warning(f'Сохранение проверки не прошло: {ex}')
        # logger.info('Сохранение проверки прошло успешно')
        # time.sleep(0.5)
        #
        #
        #
        #
        # print('Завершено')
        # time.sleep(5)
        # # Получение номера и статуса проверки кнм
        # try:
        #     self.wait_loader(self.browser.find_element(by=By.TAG_NAME, value='title'))
        #     time.sleep(5)
        #     try:
        #         knm_number = self.browser.find_element(by=By.CLASS_NAME, value='KnmHeader_Title__vO5Qr').text
        #     except:
        #         knm_number = self.browser.find_element(by=By.CLASS_NAME, value='KnmInfo_Title__LQcuP').text
        #     try:
        #         knm_status = self.browser.find_element(by=By.CLASS_NAME, value='KnmInfo_Status__3ZqpQ').text
        #     except:
        #         knm_status = self.browser.find_element(by=By.CLASS_NAME, value='KnmHeader_Status__118Oo').text
        #     result = {'number': knm_number, 'status': knm_status}
        #     print(result)
        #     logger.info('Получение номера и статуса проверки кнм')
        #
        #     return result
        # except Exception as ex:
        #     logger.warning(f'Получение номера и статуса проверки кнм не прошла: {ex}')
        #     result = {'number': None, 'status': None}
        #     print(result)
        #     return result
        # print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))
        print('пока все')
        time.sleep(10000)

    def add_knm_visited(self, knm):
        self.browser.get('https://private.proverki.gov.ru/private/knm/new')
        subject_data = list(knm[1].items())
        len_objects = len(subject_data)

        logger.info('')
        logger.info(knm[0])
        logger.info('')

        object_kinds = {
            'дошкольные образовательные организации': '174',
            'общеобразовательные организации': '175',
            'Деятельность детских лагерей на время каникул': '252',
            'школы-интернаты, специальные (коррекционные) общеобразовательные организации': '182',
            'перинатальные центры, родильные дома, родильные отделения': '162',
            'Объекты здравоохранения': '0',
            'дома (интернаты) для лиц с физическими или умственными недостатками, в том числе геронтопсихиатрические центры, психоневрологические интернаты': '161'

        }

        objects_risk = {
            'чрезвычайно высокий риск': '0',
            'высокий риск': '1',
            'значительный риск': '2',
            'средний риск': '3',
            'умеренный риск': '4',
        }

        object_danger = {
            'Первый': '0',
            'Второй': '1',
            'Третий': '2',
            'Четвертый': '3',
        }

        # принцип формирования объектов выбора: при открытии параметра выбора, добавляется див в конце страницы, чтобы узнать его айди, используй конструкцию print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML')), чтобы сделать снимок страницы в момент выпадания списка, ибо иначе его никак не захватишь, потом ища ключевую фразу( она будет последней), и смотри айдишник, на который будешь кликать

        # Наименование органа контроля

        logger.info('вставляем наименование органа контроля')
        WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div/div/main/form/div[2]/section[1]/div[2]/div/div[2]/div[1]/div/div[1]')))
        time.sleep(1)
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[2]/div/div[2]/div[1]/div/div[1]').click()

        try:
            self.browser.find_element(by=By.ID, value='react-select-2-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка наименования органа контроля не прошла: {ex}')
        logger.info('вставка наименования органа контроля прошла успешно')

        # Вид контроля

        logger.info('вставляем Вид контроля')
        WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div/div/main/form/div[2]/section[1]/div[3]/div/div[2]/div[1]/div/div[1]')))
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[3]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value='react-select-3-option-2').click()
        except Exception as ex:
            logger.warning(f'вставка вид контроля не прошла: {ex}')
        logger.info('вставка Вид контроля прошла успешно')

        # Вид КНМ

        logger.info('вставляем Вид КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[1]/div[4]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[4]/div/div[2]/div[1]/div/div[1]').click()
        try:
            self.browser.find_element(by=By.ID, value=f'react-select-4-option-6').click()
        except Exception as ex:
            logger.warning(f'вставка Вид КНМ не прошла: {ex}')
        logger.info('вставка Вид КНМ прошла успешно')

        # Характер КНМ

        self.enter_type_proverka(5)


        # Номер плана
        # time.sleep(100000)

        logger.info('вставляем Номер плана')
        self.wait_loader(WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.ID, 'planId'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.ID,
                                      value='planId').send_keys('2023050341')
        except Exception as ex:
            logger.warning(f'вставка Номер плана не прошла: {ex}')
        logger.info('вставка Номер плана прошла успешно')

        # Дата начала КНМ

        logger.info('вставляем Дата начала КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="startDateBlock"]/div[2]/div[1]/div/div/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="startDateBlock"]/div[2]/div[1]/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['начало проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                    '%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата начала КНМ не прошла: {ex}')
        logger.info('вставка Дата начала КНМ прошла успешно')

        # Дата окончания КНМ

        logger.info('вставляем Дата окончания КНМ')
        self.wait_loader(WebDriverWait(self.browser, 100).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="stopDateBlock"]/div[2]/div/div/div/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="stopDateBlock"]/div[2]/div/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['окончание проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                    '%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата окончания КНМ не прошла: {ex}')
        logger.info('вставка Дата окончания КНМ прошла успешно')

        # Cрок проведения (дней)

        logger.info('вставляем Cрок проведения (дней)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="durationDaysBlock"]/div[2]/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="durationDaysBlock"]/div[2]/input').send_keys('10')
        except Exception as ex:
            logger.warning(f'вставка Cрок проведения (дней) не прошла: {ex}')
        logger.info('вставка Cрок проведения (дней) прошла успешно')

        # Срок непосредственного взаимодействия (часов)

        logger.info('вставляем Срок непосредственного взаимодействия (часов)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="durationHoursBlock"]/div[2]/input'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="durationHoursBlock"]/div[2]/input').send_keys('80')
        except Exception as ex:
            logger.warning(f'вставка Срок непосредственного взаимодействия (часов) не прошла: {ex}')
        logger.info('вставка Срок непосредственного взаимодействия (часов) прошла успешно')

        # Основания включения в план

        logger.info('вставляем Основания включения в план')
        # жмем кнопку добавить
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.ID, "addReasonButton"))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.ID, value="addReasonButton").click()

        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[1]/div[13]/div[2]/div/div/div/div[1]/div/div[1]/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[13]/div[2]/div/div/div/div[1]/div/div[1]/div[1]').click()

        try:
            self.browser.find_element(by=By.ID, value='react-select-11-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка Основания включения в план не прошла: {ex}')
            # Не забыть вставить дату окончания последнего планового КНМ

        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="reasonsBlock"]/div[2]/div/div/div[2]/div/div[1]/div/div/input').send_keys(
                datetime.datetime.strptime(subject_data[0][1]['дата последнего планового кнм'],
                                           '%Y-%m-%d %H:%M:%S').strftime('%d.%m.%Y'))
        except Exception as ex:
            logger.warning(f'вставка Дата дата последнего планового КНМ не прошла: {ex}')

        logger.info('вставка Основания включения в план прошла успешно')

        # Наименование прокуратуры

        logger.info('вставляем Наименование прокуратуры')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[2]/div[2]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.wait_loader(self.browser.find_element(by=By.XPATH,
                                                   value='/html/body/div/div/main/form/div[2]/section[2]/div[2]/div/div[2]/div[1]/div/div[1]').click())
        try:
            self.browser.find_element(by=By.ID, value='react-select-7-option-0').click()
        except Exception as ex:
            logger.warning(f'вставка Наименование прокуратуры не прошла: {ex}')
        logger.info('вставка Наименование прокуратуры прошла успешно')

        # Основной государственный регистрационный номер (ОГРН)

        logger.info('вставляем Основной государственный регистрационный номер (ОГРН)')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '//*[@id="organizations[0].ogrn"]'))))  # используем wait_loader, так как страница запускает лоадер
        try:
            self.browser.find_element(by=By.XPATH,
                                      value='//*[@id="organizations[0].ogrn"]').send_keys(subject_data[0][1]['ОГРН'])
        except Exception as ex:
            logger.warning(f'вставка Основной государственный регистрационный номер (ОГРН) не прошла: {ex}')

        logger.info('вставка Основной государственный регистрационный номер (ОГРН) прошла успешно')

        # ждем, пока предложит варианты выпадающего списка
        try:
            WebDriverWait(self.browser, 30).until(EC.presence_of_element_located((By.ID, 'autoCompleteList')))
            self.browser.find_element(by=By.ID, value='autoCompleteList').click()
        except Exception as ex:
            logger.warning(f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) не прошел: {ex}')

        logger.info(f'выбор варианта выпадающего списка после регистрационного номера (ОГРН) прошел успешно!')

        # Добавляем количество объектов

        for sd in range(len_objects - 1):
            self.browser.find_element(by=By.ID, value='erknmObjectsAddButton').click()
            time.sleep(0.25)

        for n, subj_dat in enumerate(subject_data):

            # ставим местонахождение при необъодимости
            if n != 0:
                clipboard.copy(subj_dat[1]['адрес'])
                self.browser.find_element(by=By.XPATH,
                                          value=f'/html/body/div[1]/div/main/form/div[2]/section[3]/div[3]/div[2]/div[{n + 1}]/div/div[2]/div[1]/div/div[2]/div/textarea').send_keys(Keys.CONTROL+ 'v')
                paste = clipboard.paste
                os.system('cmd /c "echo off | clip"')
            # выбираем тип для объекта
            try:
                logger.info(f'выбираем тип для объекта {subj_dat[n]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].objectType').click()
                self.browser.find_element(by=By.ID, value=f'react-select-{12 + (n * 5)}-option-0').click()
                logger.info(f'выбор тип для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.info('')
                logger.info('foto html code')
                logger.warning(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))
                logger.warning(f'выбор тип для объекта {subj_dat[0]} не прошел: {ex}')

            time.sleep(0.25)

            # выбираем вид для объекта
            try:
                logger.info(f'выбираем вид для объекта {subj_dat[n]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].objectKind').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{13 + (n * 5)}-option-{object_kinds[subj_dat[1]["деятельность"]]}').click()
                logger.info(f'выбор вид для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор вид для объекта {subj_dat[0]} не прошел: {ex}')
            time.sleep(0.25)

            # выбираем Категория риска для объекта
            try:
                logger.info(f'выбираем Категория риска для объекта {subj_dat[n]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].riskCategory').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{15 + (n * 5)}-option-{objects_risk[subj_dat[1]["категория риска"]]}').click()
                logger.info(f'выбор Категория риска для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор Категория риска для объекта {subj_dat[0]} не прошел: {ex}')
                logger.info('')
                logger.info('foto html code')
                logger.warning(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))

            time.sleep(0.25)

            # выбираем Класс опасности для объекта
            try:
                logger.info(f'выбираем Класс опасности для объекта {subj_dat[0]}')
                self.browser.find_element(by=By.ID, value=f'objectsErknm[{n}].dangerClass').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{16 + (n * 5)}-option-{object_danger[subj_dat[1]["класс опасности"]]}').click()
                logger.info(f'выбор Класс опасности для объекта {subj_dat[0]} прошел успешно')
            except Exception as ex:
                logger.warning(f'выбор Класс опасности для объекта {subj_dat[0]} не прошел: {ex}')
            time.sleep(0.25)

        # добавляем Перечень действий, осуществляемый в рамках КНМ

        examination_ivents = ['0', '3', '4', '6', '5', '8', ]  # айдишники типов действий, осуществляемых в рамках КНМ

        try:
            for l in examination_ivents:
                self.browser.find_element(by=By.ID, value='erknmEventsAddButton').click()
                time.sleep(0.2)
            for n, lei in enumerate(examination_ivents):
                self.browser.find_element(by=By.ID, value=f'eventsErknm[{n}].type').click()
                self.browser.find_element(by=By.ID,
                                          value=f'react-select-{12 + (len_objects * 5) + n}-option-{lei}').click()
                self.browser.find_element(by=By.XPATH,
                                          value=f'//*[@id="eventsBlock"]/div[2]/div[{n + 1}]/div[2]/div/div[1]/div/div/input').send_keys(
                    datetime.datetime.strptime(subject_data[0][1]['начало проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%d.%m.%Y'))
                self.browser.find_element(by=By.XPATH,
                                          value=f'//*[@id="eventsBlock"]/div[2]/div[{n + 1}]/div[3]/div/div[1]/div/div/input').send_keys(
                    datetime.datetime.strptime(subject_data[0][1]['окончание проверки'], '%Y-%m-%d %H:%M:%S').strftime(
                        '%d.%m.%Y'))

            logger.info(f'Добавлено Перечень действий, осуществляемый в рамках КНМ прошел успешно')
        except Exception as ex:
            logger.warning(f'Добавление Перечень действий, осуществляемый в рамках КНМ не прошло: {ex}')

        # заполняем Обязательные требования, подлежащие проверке

        ND = getattr(Normative_documentation, f'_{object_kinds[subject_data[0][1]["деятельность"]]}')
        for nd in ND:
            self.browser.find_element(by=By.ID, value='erknmRequirementsAddButton').click()
            time.sleep(0.1)
        logger.info(f'Добавлено Перечень действий, осуществляемый в рамках КНМ прошел успешно')
        requirements = self.browser.find_elements(by=By.XPATH,
                                                  value='/html/body/div[1]/div/main/form/div[2]/section[4]/div[2]/table/tbody/tr')

        for n, req in enumerate(requirements):
            try:
                req.find_element(by=By.XPATH, value='./td[2]/div/div').click()

                time.sleep(1)
                formular_npa = ND[n]['ФОРМУЛИРОВКА ТРЕБОВАНИЯ']
                name_npa = ND[n]['НАИМЕНОВАНИЕ НПА']

                req.find_element(by=By.XPATH,
                                 value=f'./td[2]/div/div/div[2]//div[contains(text()="{formular_npa}")]').click()
                time.sleep(1)
                print('короткий путь')


            except Exception as ex:

                req.find_element(by=By.XPATH, value='./td[2]/div/div/div[2]//div[text()="Создать новый"]').click()
                time.sleep(1)
                clipboard.copy(ND[n]['ФОРМУЛИРОВКА ТРЕБОВАНИЯ'])

                req.find_element(by=By.XPATH, value='./td[2]/div[2]//textarea').send_keys(Keys.CONTROL +'V')
                time.sleep(2)
                clipboard.copy(ND[n]['НАИМЕНОВАНИЕ НПА'])
                req.find_element(by=By.XPATH, value='./td[3]/div[1]//textarea').send_keys(Keys.CONTROL +'V')
                print('длинный путь')
                time.sleep(2)
                clipboard.copy(ND[n]['ДАТА НПА'])
                req.find_element(by=By.XPATH, value='./td[4]/div[1]//input').send_keys(Keys.CONTROL +'V')
                time.sleep(1)

        # отметка об использовании проверочного листа

        logger.info('Нажимаем отметка об использовании проверочного листа')
        while True:
            try:
                self.wait_loader(self.browser.find_element(by=By.ID, value='isChecklistsUsed').click())
            except Exception as ex:
                logger.warning(f'Нажатие отметки об использовании проверочного листа не прошла: {ex}')
            logger.info('Нажатие отметки об использовании проверочного листа прошла успешно')
            try:
                self.browser.find_element(by=By.XPATH, value='//*[@id="checklistsBlock"]/div/button')
                break
            except:
                continue
            time.sleep(1)

        # Добавляем проверочный лист

        PL = getattr(Proverochnii_list, f'_{object_kinds[subject_data[0][1]["деятельность"]]}')
        print(len(PL))
        for every_pl in PL:
            logger.info('Добавляем проверочный лист')
            try:
                self.browser.find_element(by=By.XPATH, value='//*[@id="checklistsBlock"]/div/button').click()

            except Exception as ex:
                logger.warning(f'Добавляем проверочный лист не прошла: {ex}')
            logger.info('Добавляем проверочный лист прошла успешно')
            time.sleep(0.5)

        #  выбираем проверочный лист
        logger.info('выбираем проверочный лист')
        for n, pl in enumerate(PL):
            try:
                self.browser.find_element(By.XPATH,
                                          value=f'/html/body/div[1]/div/main/form/div[2]/section[4]/div[6]/div/ul/li[{n + 1}]/div/div[2]/div[2]/div[2]/div/div/div/div[1]').click()
                self.browser.find_element(by=By.XPATH,
                                          value=f'/html/body/div[1]/div/main/form/div[2]/section[4]/div[6]/div/ul/li[{n + 1}]/div/div[2]/div[2]/div[2]/div/div/div/div[1]//div[text()="{pl}"]').click()
                # logger.info(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))

            except Exception as ex:
                logger.warning(f'Выбор проверочных листов не прошел: {ex}')
                print(ex)
            time.sleep(2)
        logger.info('Выбор всех проверочных листов успешно завершен')

        # Добавляем места проведения контрольного (надзорного) мероприятия

        for o in subject_data:
            logger.info('Добавляем места проведения контрольного (надзорного) мероприятия')
            try:
                self.browser.find_element(by=By.ID, value='erknmPlacesAddButton').click()

            except Exception as ex:
                logger.warning(f'Добавление места проведения контрольного (надзорного) мероприятия не прошла: {ex}')
        logger.info('Добавление места проведения контрольного (надзорного) мероприятия прошла успешно')
        time.sleep(0.5)

        # Вставляем места проведения контрольного (надзорного) мероприятия
        for n, obj in enumerate(subject_data):
            logger.info('Вставляем места проведения контрольного (надзорного) мероприятия')
            try:
                self.browser.find_element(by=By.XPATH,
                                          value=f'/html/body/div/div/main/form/div[2]/section[4]/div[7]/div[2]/div[{n + 1}]/div/div/textarea').send_keys(
                    str(obj[1]['адрес']))

            except Exception as ex:
                logger.warning(f'Вставляем места проведения контрольного (надзорного) мероприятия не прошла: {ex}')
        logger.info('Вставляем места проведения контрольного (надзорного) мероприятия прошла успешно')

        # os.system('msg * Программа закончила, смотри сам что получилось')
        # time.sleep(100000)
        return self.save_proverka()


    def enter_type_proverka(self, number:int):
        logger.info('вставляем Характер КНМ')
        self.wait_loader(WebDriverWait(self.browser, 20).until(EC.presence_of_element_located(
            (By.XPATH,
             '/html/body/div/div/main/form/div[2]/section[1]/div[5]/div/div[2]/div[1]/div/div[1]'))))  # используем wait_loader, так как страница запускает лоадер
        self.browser.find_element(by=By.XPATH,
                                  value='/html/body/div/div/main/form/div[2]/section[1]/div[5]/div/div[2]/div[1]/div/div[1]').click()
        try:
            time.sleep(2)
            WebDriverWait(self.browser, 20).until(EC.presence_of_element_located((By.ID, f'react-select-{number}-option-1'))).click()
        except Exception as ex:
            logger.warning(f'вставка Характер КНМ не прошла: {ex}')
        logger.info('вставка Характер КНМ прошла успешно')



    def save_proverka(self):

        # сохраняем проверку

        try:
            self.wait_loader(self.browser.find_element(by=By.ID, value='saveButton').click())

        except Exception as ex:
            logger.warning(f'Сохранение проверки не прошло: {ex}')
        logger.info('Сохранение проверки прошло успешно')
        time.sleep(0.5)

        print('Завершено')
        time.sleep(5)
        # Получение номера и статуса проверки кнм
        WebDriverWait(self.browser, 100).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="root"]/div/header/div/div[1]/div')))
        try:
            self.wait_loader(self.browser.find_element(by=By.TAG_NAME, value='title'))
            time.sleep(5)
            try:
                knm_number = self.browser.find_element(by=By.CLASS_NAME, value='_Title_13www_8').text
            except:
                knm_number = self.browser.find_element(by=By.CLASS_NAME, value='_Title_13www_8').text
            try:
                knm_status = self.browser.find_element(by=By.CLASS_NAME, value='_Title_5dm4o_45').text
            except:
                try:
                    knm_status = self.browser.find_element(by=By.CLASS_NAME, value='_Title_13www_8').text
                except:
                    knm_status = 'не удалось подгрузить статус'

            result = {'number': knm_number, 'status': knm_status}
            print(result)
            logger.info('Получение номера и статуса проверки кнм')

            return result
        except Exception as ex:
            logger.warning(f'Получение номера и статуса проверки кнм не прошла: {ex}')


            result = {'number': None, 'status': None}

            print(result)
            return result

        # print(self.browser.find_element(by=By.TAG_NAME, value='html').get_attribute('innerHTML'))

    def wait_loader(self, func):
        while True:
            try:
                time.sleep(0.25)
                bool(self.browser.find_element(by=By.XPATH, value='//*[@id="root"]/div[2]'))
                time.sleep(2)
            except:
                break
        func


if __name__ == '__main__':
    erknm()
