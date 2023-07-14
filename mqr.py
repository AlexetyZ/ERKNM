import json
import os
from datetime import datetime

import openpyxl


class QuiqlyReport:
    def __init__(self, path):
        with open(path, 'r') as file:
            self.allInspection = json.load(file)
            self.inspections = self.allInspection

    def reset_inspection_filter(self):
        self.inspections = self.allInspection

    def count_inspections(self):
        return len(self.inspections)


    def simply_clever(self, start, stop):
        main_format_date = '%d.%m.%Y'
        start_limit = datetime.strptime(start, main_format_date)
        stop_limit = datetime.strptime(stop, main_format_date)
        return self.inspections.count(start_limit < datetime.strptime('startDate', main_format_date) < stop_limit)

    def report_planed_and_conducted_by_regions(self, start: str, stop: str):
        result = {}

        main_format_date = '%d.%m.%Y'
        start_limit = datetime.strptime(start, main_format_date)
        stop_limit = datetime.strptime(stop, main_format_date)
        for inspection in self.inspections:
            inspection_date = datetime.strptime(lambda inspection: inspection['startDate'], main_format_date)
            if start_limit < inspection_date < stop_limit:

                if inspection['controllingOrganization'] in result:
                    pass

    def countAllObjects(self, planned=True, outplanned=False, pm=True):
        object_count = 0

        for inspection in self.inspections:
            if planned:
                if inspection['planId']:
                    object_count += len(inspection['addresses'])

            if outplanned:
                if not inspection['planId']:
                    object_count += len(inspection['addresses'])
            if pm:
                object_count += len(inspection['addresses'])
        return object_count

    def structureByKindControl(self, inspections) -> dict:
        result = {}

        seb = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}
        zpp = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}
        zashDet = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}
        license_viz = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}
        license_iii = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}

        for inspection in inspections:
            if inspection['status'] == 'Завершено':
                try:
                    stopMonth = str(inspection['stopDate']).split('.')[1]
                except:
                    stopMonth = str(inspection['startDate']).split('.')[1]
                if int(stopMonth) > int(datetime.now().month):
                    print(inspection)
                    continue

                if inspection['supervisionTypeId'] == '004':
                    seb[stopMonth] += 1
                if inspection['supervisionTypeId'] == '079':
                    zpp[stopMonth] += 1
                if inspection['supervisionTypeId'] == '130':
                    zashDet[stopMonth] += 1
                if inspection['supervisionTypeId'] == '032':
                    license_viz[stopMonth] += 1
                if inspection['supervisionTypeId'] == '031':
                    license_iii[stopMonth] += 1
        result['виды надзора\месяцы'] = {'01': 'январь', '02': "февраль", '03': "март", '04': "апрель", '05': "май", '06': "июнь", '07': "июль", '08': "август", '09': "сентябрь", '10': "октябрь", '11': "ноябрь", '12': "декабрь"}
        result['СЭБ'] = seb
        result['ЗПП'] = zpp
        result['Защита детей'] = zashDet
        result['ВИЗ'] = license_viz
        result['ИИИ'] = license_iii

        return result

    def pack_to_exel_line_model(self, dictionary, path_file: str = None):
        wb = openpyxl.Workbook()
        sh = wb.worksheets[0]
        for kind, values in dictionary.items():
            sh.append((kind, *values.values()))
        if path_file:
            wb.save(path_file)
        else:
            wb.save('report_line_model.xlsx')

    def filter(self, **conditions):
        filtered = []
        for inspection in self.inspections:
            maybe = True
            for key, value in conditions.items():
                if inspection[key] != value:
                    maybe = False
                    break
            if maybe:
                filtered.append(inspection)
        self.inspections = filtered

    def countRiskCategory(self, inspections, planned=True, outplanned=False, ):
        extraHigh = 0
        high = 0

        for inspection in inspections:
            if planned:
                for risk in inspection['riskCategory']:
                    if risk == 'чрезвычайно высокий риск':
                        extraHigh += 1
                    elif risk == 'высокий риск':
                        high += 1

            if outplanned:
                for risk in inspection['riskCategory']:
                    if risk == 'чрезвычайно высокий риск':
                        extraHigh += 1
                    elif risk == 'высокий риск':
                        high += 1

        return {'чрезвычайно высокий риск': extraHigh, "высокий риск": high}

def count_pm_object_by_tu_and_form_to_xl():
    qr = QuiqlyReport('pm_2023.json')
    xl_path = "C:\\Users\zaitsev_ad\Desktop\сведения из базы данных.xlsx"
    wb = openpyxl.load_workbook(xl_path)
    sh = wb.worksheets[0]
    cells = sh['A:A']
    for n, cell in enumerate(cells):
        if cell.value:
            qr.filter(
                controllingOrganizationId=cell.value,
                status="Завершено"
            )
            object_count = qr.countAllObjects(planned=False, outplanned=False, pm=True)
            sh.cell(row=n + 1, column=3, value=object_count)
            qr.reset_inspection_filter()
    wb.save(xl_path)


def kind_controll_structure_and_form_to_xl():
    qr = QuiqlyReport('pm_2023.json')
    dictionary = qr.structureByKindControl(qr.inspections)
    qr.pack_to_exel_line_model(dictionary)


def kind_controll_structure_by_tu_and_form_to_xl():
    qr = QuiqlyReport('Plan_knm_full_2023.json')
    xl_path = "C:\\Users\zaitsev_ad\Desktop\сведения из базы данных.xlsx"
    dir_path = "C:\\Users\zaitsev_ad\Desktop\отчет про регионам"
    if not os.path.exists(dir_path):
        os.mkdir(dir_path)
    main_wb = openpyxl.load_workbook(xl_path)
    main_sh = main_wb.worksheets[0]

    for row in main_sh.iter_rows(min_col=1, max_col=3, values_only=True):
        if row:
            qr.filter(
                controllingOrganizationId=row[0],
                status="Завершено"
            )
            dictionary = qr.structureByKindControl(qr.inspections)
            qr.pack_to_exel_line_model(dictionary, path_file=f'{dir_path}\{row[1]}.xlsx')
            qr.reset_inspection_filter()


if __name__ == '__main__':
    print(datetime.now())
    kind_controll_structure_by_tu_and_form_to_xl()
    print(datetime.now())




    # QuiqlyReport().pack_to_exel_line_model(responce)

