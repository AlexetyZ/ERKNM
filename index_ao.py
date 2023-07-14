import openpyxl
from pathlib import Path

class Merge:
    def __init__(self, path_1, path_2):
        self.path_1 = path_1
        self.path_2 = path_2
        self.wb1 = openpyxl.load_workbook(path_1)
        self.wb2 = openpyxl.load_workbook(path_2)

        self.sh1 = self.wb1.worksheets[0]
        self.sh2 = self.wb2.worksheets[0]
        self.get_reserve()

    def activate(self, column):
        for index, sh1_row in enumerate(self.sh1.iter_rows(min_row=6, min_col=2, values_only=True)):
            for sh2_row in self.sh2.iter_rows(min_row=2, min_col=1, values_only=True):
                if sh2_row[0] == sh1_row[0]:
                    self.sh1.cell(row=index+6, column=column, value=sh2_row[1])

        self.wb1.save(self.path_1)

    def get_reserve(self):
        self.wb1.save(f"{Path(self.path_1).parent}\\копия-{Path(self.path_1).name}")



if __name__ == '__main__':
    # path_1 = "C:\\Users\zaitsev_ad\Documents\отчет ИАД\Приложение Индекс (1).xlsx"
    # path_2 = "C:\\Users\zaitsev_ad\Documents\отчет ИАД\количество_объектов_по_регионам_2023-04-12T12_51_12.935316Z.xlsx"
    # Merge(path_1, path_2).activate(9)


    str1 = 'I want to become a Python developer'
    print(str1.split(' want '))
