import os.path
from sys import argv
from tqdm import tqdm
from pprint import pprint
import Dictionary as d
from private_config import default_path_to_save_result
from mongo_database import convertForsaving, unpac_idAggregation

# *group_kinds['Деятельность в сфере здравоохранения'],
# *group_kinds['Детские лагеря с дневным пребыванием'],
# *group_kinds['Деятельность общеобразовательных организаций'],
# *group_kinds['Деятельность дошкольных образовательных организаций'],
# *group_kinds['Объекты промышленности и транспорта'],
# *group_kinds['Деятельность водоснабжения и водоотведения'],
# *group_kinds['Коммунальное обслуживание'],
# *group_kinds['Деятельность по предоставлению персональных услуг'],
# *group_kinds['Деятельность гостиниц и прочих мест для временного проживания'],
# *group_kinds['Предоставление социальных услуг'],
# *group_kinds['Деятельность детских и подростковых организаций, образования, в том числе дополнительного образования'],
# *group_kinds['Деятельность в области обращения с отходами'],
# *group_kinds['Работы с микроорганизмами'],
# *group_kinds['Организации для детей-сирот'],
# *group_kinds['Деятельность по организации отдыха и развлечений, культуры и спорта'],
# *group_kinds['Профессиональные образовательные организации'],
# *group_kinds['Учреждения высшего профессионального образования'],
# *group_kinds['Деятельность организаторов детского питания'],
# *group_kinds['Торговля пищевыми продуктами'],
# *group_kinds['Производство пищевых продуктов'],
# *group_kinds['Общественное питание населения'],
# *group_kinds['Продукция'],




group_kinds = {
    'Деятельность в сфере здравоохранения': d.health_care_kinds,
    'Детские лагеря с дневным пребыванием': d.child_camps_kinds,
    'Деятельность организаторов детского питания': d.children_meal_kinds,
    'Деятельность общеобразовательных организаций': d.school_kinds,
    'Деятельность дошкольных образовательных организаций': d.preschool_kinds,
    'Торговля пищевыми продуктами': d.food_store_kind,
    'Объекты промышленности и транспорта': d.industry_kinds,
    'Производство пищевых продуктов': d.food_production,
    'Общественное питание населения': d.public_catering_kind,
    'Деятельность водоснабжения и водоотведения': d.water_supply_kinds,
    'Коммунальное обслуживание': d.communal_services,
    'Деятельность по предоставлению персональных услуг': d.personal_services,
    'Деятельность гостиниц и прочих мест для временного проживания': d.motels,
    'Предоставление социальных услуг': d.all_social_services,

    'Деятельность детских и подростковых организаций, образования, в том числе дополнительного образования': [
        *d.child_organization,
        *d.education,
        *d.other_child_organization,
        *d.addictional_education
    ],
    'Деятельность в области обращения с отходами': d.trash,
    'Работы с микроорганизмами': d.patogens_work,
    'Организации для детей-сирот': d.orphans_child,
    'Деятельность по организации отдыха и развлечений, культуры и спорта': d.relax_animation_sport_culture,
    'Работы с источниками ионизирующего излучения': d.ionizing_radiation,
    'Профессиональные образовательные организации': d.professional_education,
    'Учреждения высшего профессионального образования': d.high_education,
    'Деятельность в области связи': d.connection,
    'Продукция': d.production
}


def load_knm(year):
    from ERKNM_http import Erknm
    Erknm().get_all_knm_for_a_year(year=year)


def load_pm(year):
    from ERKNM_http import Erknm
    Erknm().get_all_pm_for_a_year(year=year)


def reportByObjects(pathFile, filters: str):
    from mongo_database import WorkMongo
    import xl
    import json

    filters = json.loads(filters.replace("'", '"')) if filters else {}
    wm = WorkMongo()
    _objects = wm.reportByObjects(**filters)
    results = convertForsaving(list(_objects))
    xl.writeResultsInXL(title=results[0], results=results[1], pathFile=pathFile)


def mergeTu(path, col_number):
    import xl
    tu_list = xl.bringCol(path, col_number)
    return xl.merge_tu(tu_list)


def reportByRisks(pathFile, filters: str):
    from mongo_database import WorkMongo
    import xl
    import json

    filters = json.loads(filters.replace("'", '"')) if filters else {}
    wm = WorkMongo()
    _objects = wm.reportByRisk(**filters)
    results = convertForsaving(list(_objects))
    xl.writeResultsInXL(title=results[0], results=results[1], pathFile=pathFile)


def reportByFreeCommand(pathFile, command):
    from mongo_database import WorkMongo
    import xl
    import json

    filters = json.loads(command.replace("'", '"')) if command else {}
    wm = WorkMongo()
    _objects = wm.free_command(filters)
    results = convertForsaving(list(_objects))
    xl.writeResultsInXL(title=results[0], results=results[1], pathFile=pathFile)


def reportByProsecutorComments(pathFile):
    from mongo_database import WorkMongo
    import xl
    import json

    wm = WorkMongo()
    _objects = wm.reportFromDeniedKNMComment()
    results = convertForsaving(list(_objects))
    xl.writeResultsInXL(title=results[0], results=results[1], pathFile=pathFile)
    return 'Причины исключений успешно выгружены!'


def reportByDienedObjects(countBy):
    from mongo_database import WorkMongo
    import xl
    from pathlib import Path
    import os
    import json
    from datetime import datetime
    import Dictionary as d
    from private_config import dayliProcessFile

    group_kinds = {
        'Деятельность в сфере здравоохранения': d.health_care_kinds,
        'Детские лагеря с дневным пребыванием': d.child_camps_kinds,
        'Деятельность организаторов детского питания': d.children_meal_kinds,
        'Деятельность общеобразовательных организаций': d.school_kinds,
        'Деятельность дошкольных образовательных организаций': d.preschool_kinds,
        'Торговля пищевыми продуктами': d.food_store_kind,
        'Объекты промышленности и транспорта': d.industry_kinds,
        'Производство пищевых продуктов': d.food_production,
        'Общественное питание населения': d.public_catering_kind,
        'Деятельность водоснабжения и водоотведения': d.water_supply_kinds,
        'Коммунальное обслуживание': d.communal_services,
        'Деятельность по предоставлению персональных услуг': d.personal_services,
        'Деятельность гостиниц и прочих мест для временного проживания': d.motels,
        'Предоставление социальных услуг': d.all_social_services,

        'Деятельность детских и подростковых организаций, образования, в том числе дополнительного образования': [
            *d.child_organization,
            *d.education,
            *d.other_child_organization,
            *d.addictional_education
        ],
        'Деятельность в области обращения с отходами': d.trash,
        'Работы с микроорганизмами': d.patogens_work,
        'Организации для детей-сирот': d.orphans_child,
        'Деятельность по организации отдыха и развлечений, культуры и спорта': d.relax_animation_sport_culture,
        'Работы с источниками ионизирующего излучения': d.ionizing_radiation,
        'Профессиональные образовательные организации': d.professional_education,
        'Учреждения высшего профессионального образования': d.high_education,
        'Деятельность в области связи': d.connection,
        'Продукция': d.production
    }

    t_data = datetime.now().strftime('%d.%m.%Y')

    wm = WorkMongo()
    _tus_kinds_counts = wm.reportFromDeniedKNMObjectCategory() if countBy == 'objects' else wm.reportFromDeniedKNMObjectCategoryKNM()

    results = {'Всего': {'Общее': 0}}
    value_list = []
    # print(results['Всего']['Общее'])
    otherKinds = []
    for tu_object_count in _tus_kinds_counts:
        recordId = tu_object_count['_id']
        count = tu_object_count['totalCount']
        tu = recordId['tu']
        kind = recordId['kind']
        isOther = True
        for groupName, groupRefs in group_kinds.items():
            if kind in groupRefs:
                kind = groupName
                isOther = False
        if isOther:
            otherKinds.append(kind)
            kind = 'Прочие виды деятельности'

        if tu in results:
            # print(results[tu])
            if kind in results[tu]:
                results[tu][kind] += count
            else:
                results[tu][kind] = count
            results[tu]['Общее'] += count

        else:
            results[tu] = {'Общее': count, kind: count}
        if kind in results['Всего']:
            results['Всего'][kind] += count
        else:
            results['Всего'][kind] = count
        # pprint(results)
        results['Всего']['Общее'] += count

    results = dict(sorted(results.items(), key=lambda x: x[1]['Общее'], reverse=True))
    # pprint(results)
    results['Всего'] = dict((sorted(results['Всего'].items(), key=lambda x: x[1], reverse=True)))

    value_list.append(['№', 'ТУ', *[key for key in results['Всего'].keys()]])
    # for main_kind in results['Всего'].keys():+

    # print(list(results['Всего'].keys()))
    for number, (tu, kinds) in enumerate(results.items()):
        # print(kinds)
        value_list.append([number if number else '', tu,
                           *[kinds[val] if val in kinds else 0 for val in list(results['Всего'].keys())]])

    pathFile = dayliProcessFile(f"отчет по отклоненным объектам {t_data} {countBy}.xlsx")

    xl.writeResultsInXL(results=value_list,
                        title=f'Структура по видам деятельности объектов, исключенных из плана прокуратурой на {t_data}' if countBy == 'objects' else f'Структура КНМ по видам деятельности объектов, исключенных из плана прокуратурой на {t_data}',
                        pathFile=pathFile)
    xl.writeResultsInXL(results=[kind for kind in set(otherKinds)], title=f'Прочие виды деятельности',
                        pathFile=pathFile, sheetTitle='прочие', sheetIndex=1)
    dienedObjectsStyles(pathFile)
    pathFileStr = str(pathFile).replace("\\\\", "\\")
    # pprint(otherKinds)
    return f'Отчет готов! Сохранен в {pathFileStr}'


def reportByAccepedObjects(countBy):
    from mongo_database import WorkMongo
    import xl
    from pathlib import Path
    import os
    import json
    from datetime import datetime
    import Dictionary as d
    from private_config import dayliProcessFile

    group_kinds = {
        'Деятельность в сфере здравоохранения': d.health_care_kinds,
        'Детские лагеря с дневным пребыванием': d.child_camps_kinds,
        'Деятельность организаторов детского питания': d.children_meal_kinds,
        'Деятельность общеобразовательных организаций': d.school_kinds,
        'Деятельность дошкольных образовательных организаций': d.preschool_kinds,
        'Торговля пищевыми продуктами': d.food_store_kind,
        'Объекты промышленности и транспорта': d.industry_kinds,
        'Производство пищевых продуктов': d.food_production,
        'Общественное питание населения': d.public_catering_kind,
        'Деятельность водоснабжения и водоотведения': d.water_supply_kinds,
        'Коммунальное обслуживание': d.communal_services,
        'Деятельность по предоставлению персональных услуг': d.personal_services,
        'Деятельность гостиниц и прочих мест для временного проживания': d.motels,
        'Предоставление социальных услуг': d.all_social_services,

        'Деятельность детских и подростковых организаций, образования, в том числе дополнительного образования': [
            *d.child_organization,
            *d.education,
            *d.other_child_organization,
            *d.addictional_education
        ],
        'Деятельность в области обращения с отходами': d.trash,
        'Работы с микроорганизмами': d.patogens_work,
        'Организации для детей-сирот': d.orphans_child,
        'Деятельность по организации отдыха и развлечений, культуры и спорта': d.relax_animation_sport_culture,
        'Работы с источниками ионизирующего излучения': d.ionizing_radiation,
        'Профессиональные образовательные организации': d.professional_education,
        'Учреждения высшего профессионального образования': d.high_education,
        'Деятельность в области связи': d.connection,
        'Продукция': d.production
    }

    t_data = datetime.now().strftime('%d.%m.%Y')

    wm = WorkMongo()
    _tus_kinds_counts = wm.reportFromAcceptKNMObjectCategory()

    results = {'Всего': {'Общее': 0}}
    value_list = []
    # print(results['Всего']['Общее'])
    otherKinds = []
    for tu_object_count in _tus_kinds_counts:
        recordId = tu_object_count['_id']
        count = tu_object_count['totalCount']
        tu = recordId['tu']
        kind = recordId['kind']
        isOther = True
        for groupName, groupRefs in group_kinds.items():
            if kind in groupRefs:
                kind = groupName
                isOther = False
        if isOther:
            otherKinds.append(kind)
            kind = 'Прочие виды деятельности'

        if tu in results:
            # print(results[tu])
            if kind in results[tu]:
                results[tu][kind] += count
            else:
                results[tu][kind] = count
            results[tu]['Общее'] += count

        else:
            results[tu] = {'Общее': count, kind: count}
        if kind in results['Всего']:
            results['Всего'][kind] += count
        else:
            results['Всего'][kind] = count
        # pprint(results)
        results['Всего']['Общее'] += count

    results = dict(sorted(results.items(), key=lambda x: x[1]['Общее'], reverse=True))
    # pprint(results)
    results['Всего'] = dict((sorted(results['Всего'].items(), key=lambda x: x[1], reverse=True)))

    value_list.append(['№', 'ТУ', *[key for key in results['Всего'].keys()]])
    # for main_kind in results['Всего'].keys():+

    # print(list(results['Всего'].keys()))
    for number, (tu, kinds) in enumerate(results.items()):
        # print(kinds)
        value_list.append([number if number else '', tu,
                           *[kinds[val] if val in kinds else 0 for val in list(results['Всего'].keys())]])

    pathFile = dayliProcessFile(f"отчет по согласованным объектам {t_data} {countBy}.xlsx")

    xl.writeResultsInXL(results=value_list,
                        title=f'Структура по видам деятельности объектов, согласованных прокуратурой на {t_data}' if countBy == 'objects' else f'Структура КНМ по видам деятельности объектов, согласованных прокуратурой на {t_data}',
                        pathFile=pathFile)
    xl.writeResultsInXL(results=[kind for kind in set(otherKinds)], title=f'Прочие виды деятельности',
                        pathFile=pathFile, sheetTitle='прочие', sheetIndex=1)
    dienedObjectsStyles(pathFile)
    pathFileStr = str(pathFile).replace("\\\\", "\\")
    # pprint(otherKinds)
    return f'Отчет готов! Сохранен в {pathFileStr}'


def dienedObjectsStyles(pathFile):
    import xl
    from openpyxl.styles import Font
    someCols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
     'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
    colsWidth = {
        'A': 4,
        'B': 28.83,
        **{letter: 8.85 for letter in someCols}
    }
    merge_cells = [
        'A1:AB1',
    ]
    positional = {
        'A1:I1': 'center',
        'C2:ABmax': 'center',
        'A4:Amax': 'center'
    }
    cellsWrap = [
        'B2:Bmax',
        'B2:AB2'
    ]

    formatNumber = {
        # 'E3:Emax': '0.0%',
        # 'I3:Imax': '0.0%',

    }

    borderForAllCellInRange = [
        f'{letter}1:{letter}max' for letter in ['A', 'B', *someCols]
    ]
    printArea = 'A1:ABmax'
    printRows = '1:3'
    fitToWidht = 1
    fitToHeight = 5

    style1 = Font(bold=True, sz=11, name='Times New Roman')
    style2 = Font(sz=11, name='Times New Roman')
    fonts = {
        'A:A': style1,
        'B2:AB4': style1,
        'B2:Cmax': style1,
        'D5:Imax': style2
    }
    # Worksheet.PAPERSIZE_A3 = '8'.
    # Worksheet.PAPERSIZE_A4 = '9'.
    # Worksheet.PAPERSIZE_A4_SMALL = '10'.
    # Worksheet.PAPERSIZE_A5 = '11'.
    # Worksheet.PAPERSIZE_EXECUTIVE = '7'.
    # Worksheet.PAPERSIZE_LEDGER = '4'.
    # Worksheet.PAPERSIZE_LEGAL = '5'.
    # Worksheet.PAPERSIZE_LETTER = '1'.
    # Worksheet.PAPERSIZE_LETTER_SMALL = '2'.
    # Worksheet.PAPERSIZE_STATEMENT = '6'.
    # Worksheet.PAPERSIZE_TABLOID = '3'.

    paperSize = 8

    xl.formatFile(
        pathFile=pathFile,
        columsWidth=colsWidth,
        mergeCells=merge_cells,
        cellsWrap=cellsWrap,
        positionText=positional,
        formatNumber=formatNumber,
        borderForAllCellInRange=borderForAllCellInRange,
        paperSize=paperSize,
        printArea=printArea,
        printRows=printRows,
        fitToWidht=fitToWidht,
        fitToHeight=fitToHeight,
        fonts=fonts,
        sheetIndex=0
    )

def dailyAggreedProcessStyles(pathFile):
    import xl
    from openpyxl.styles import Font
    colsWidth = {
        'A': 4,
        'B': 32.67,
        'C': 17.17,
        'D': 17.17,
        'E': 18.33,
        'F': 17.67,
        'G': 17.67,
        'H': 27.83,
        'I': 19.17
    }
    merge_cells = [
        'A1:I1',
    ]
    positional = {
        'A1:I1': 'center',
        'C2:Imax': 'center',
        'A4:Amax': 'center'
    }
    alignmentFor = 'A:I'

    formatNumber = {
        'E3:Emax': '0.0%',
        'I3:Imax': '0.0%',

    }

    borderForAllCellInRange = [
        f'{letter}1:{letter}max' for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    ]
    printArea = 'A1:Imax'
    printRows = '1:3'
    fitToWidht = 1
    fitToHeight = 5

    style1 = Font(bold=True, sz=14, name='Times New Roman')
    style2 = Font(sz=14, name='Times New Roman')
    fonts = {
        'A:A': style1,
        'B2:I3': style1,
        'B2:Cmax': style1,
        'D4:Imax': style2
    }
    # Worksheet.PAPERSIZE_A3 = '8'.
    # Worksheet.PAPERSIZE_A4 = '9'.
    # Worksheet.PAPERSIZE_A4_SMALL = '10'.
    # Worksheet.PAPERSIZE_A5 = '11'.
    # Worksheet.PAPERSIZE_EXECUTIVE = '7'.
    # Worksheet.PAPERSIZE_LEDGER = '4'.
    # Worksheet.PAPERSIZE_LEGAL = '5'.
    # Worksheet.PAPERSIZE_LETTER = '1'.
    # Worksheet.PAPERSIZE_LETTER_SMALL = '2'.
    # Worksheet.PAPERSIZE_STATEMENT = '6'.
    # Worksheet.PAPERSIZE_TABLOID = '3'.

    paperSize = 8

    xl.formatFile(
        pathFile=pathFile,
        columsWidth=colsWidth,
        mergeCells=merge_cells,
        cellsWrap=alignmentFor,
        positionText=positional,
        formatNumber=formatNumber,
        borderForAllCellInRange=borderForAllCellInRange,
        paperSize=paperSize,
        printArea=printArea,
        printRows=printRows,
        fitToWidht=fitToWidht,
        fitToHeight=fitToHeight,
        fonts=fonts,
        sheetIndex=1
    )


def reportByAggreedProcess(pathFile, today, count_by='knm'):
    from mongo_database import WorkMongo
    import xl
    from pathlib import Path
    import os
    import json

    parentDir = Path(pathFile).parent
    previewsResultFile = os.path.join(parentDir, f'preview_results {count_by}.json')
    value_list = []
    results = {}
    statusses = []
    wm = WorkMongo()
    if count_by == 'knm':
        _objects = wm.reportByAggreedProcessKNM()
    elif count_by == 'objects':
        _objects = wm.reportByAggreedProcessObjects()
    else:
        return 'Не удалось распознать дополнительный аттрибут, по которому будет подсчет!'
    for _object in _objects:
        id_part = _object['_id']
        name = id_part['tu']
        status = id_part['status']
        count = _object['totalCount']
        statusses.append(status)
        if name in results:
            results[name][status] = count  # ТУ с определенным статусом попадается только 1 раз
        else:
            results[name] = {status: count}
    statusses = set(statusses)
    onApplyKinds = ['Есть замечания', 'На согласовании', 'Исключена', 'Готово к согласованию', 'Ожидает проведения',
                    'Не согласована']
    registredExcludedKnm = []
    totalExcluded = {'total': 0, 'excluded': 0, 'persentExcludedTotal': 0, 'onApply': 0, 'applied': 0,
                     'increaseExcluded': 0,
                     'persentExcludedOnDate': 0}
    if not os.path.exists(previewsResultFile):

        for tu, values in results.items():
            tuIncreaseExcluded = 0
            v = []
            onApplyCount = 0
            for s in statusses:
                if s in values:
                    if s in onApplyKinds:
                        onApplyCount += values[s]
                    v.append(values[s])

                    v.append(0)
                else:
                    v.append(0)
                    v.append(0)
            totalExcluded['total'] += onApplyCount
            isNotes = values['Есть замечания'] if 'Есть замечания' in values else 0
            waitForStarting = values['Ожидает проведения'] if 'Ожидает проведения' in values else 0
            totalExcluded['applied'] += isNotes + waitForStarting
            onApply = values['На согласовании'] if 'На согласовании' in values else 0
            totalExcluded['onApply'] += onApply
            if 'Исключена' in values:
                excluded = values['Исключена']
                if excluded:
                    existExcluded = [tu, onApplyCount, excluded]
                    totalExcluded['excluded'] += excluded
                    persentExcludedTotal = excluded / onApplyCount
                    existExcluded.append(persentExcludedTotal)
                    existExcluded.append(onApply)
                    applied = isNotes + waitForStarting
                    existExcluded.append(applied)
                    existExcluded.append(tuIncreaseExcluded)
                    totalExcluded['increaseExcluded'] += tuIncreaseExcluded
                    persentExcludedOnDate = tuIncreaseExcluded / onApplyCount
                    existExcluded.append(persentExcludedOnDate)
                    registredExcludedKnm.append(existExcluded)

            value_list.append([tu, *v, onApplyCount])

    else:
        with open(previewsResultFile, 'r', encoding='utf-8') as previewsFile:
            previewsResult = json.load(previewsFile)

        for tu, values in results.items():
            tuIncreaseExcluded = 0
            v = []
            onApplyCount = 0

            for s in statusses:
                if s in values:
                    if s in onApplyKinds:
                        onApplyCount += values[s]

                    v.append(values[s])

                    tuIncrease = values[s] - previewsResult[tu][s] if s in previewsResult[tu] else values[s]
                    if s == "Исключена":
                        tuIncreaseExcluded = tuIncrease

                    v.append(tuIncrease)
                else:
                    v.append(0)
                    v.append(0)
            totalExcluded['total'] += onApplyCount
            isNotes = values['Есть замечания'] if 'Есть замечания' in values else 0
            waitForStarting = values['Ожидает проведения'] if 'Ожидает проведения' in values else 0
            totalExcluded['applied'] += isNotes + waitForStarting
            onApply = values['На согласовании'] if 'На согласовании' in values else 0
            totalExcluded['onApply'] += onApply

            if 'Исключена' in values:
                excluded = values['Исключена']
                if excluded:
                    existExcluded = [tu, onApplyCount, excluded]
                    totalExcluded['excluded'] += excluded
                    persentExcludedTotal = excluded / onApplyCount
                    existExcluded.append(persentExcludedTotal)
                    existExcluded.append(onApply)
                    applied = isNotes + waitForStarting
                    existExcluded.append(applied)
                    existExcluded.append(tuIncreaseExcluded)
                    totalExcluded['increaseExcluded'] += tuIncreaseExcluded
                    persentExcludedOnDate = tuIncreaseExcluded / onApplyCount
                    existExcluded.append(persentExcludedOnDate)
                    registredExcludedKnm.append(existExcluded)
            value_list.append([tu, *v, onApplyCount])

    excludedReportTitle = [
        '',
        'ТУ',
        "Всего КНМ в проекте плана" if count_by == 'knm' else "Всего объектов в проекте плана",
        "Исключено прокуратурой",
        '% исключенных всего',
        'Осталось на согласовании',
        'Согласовано',
        f'Прирост исключенных за {today}',
        f'% исключенных всего за {today}'
    ]

    totalExcluded['persentExcludedTotal'] = totalExcluded['excluded'] / totalExcluded['total']
    totalExcluded['persentExcludedOnDate'] = totalExcluded['increaseExcluded'] / totalExcluded['total']
    registredExcludedKnm = sorted(registredExcludedKnm, key=lambda x: x[3], reverse=True)
    registredExcludedKnm = [[number+1, *region] for number, region in enumerate(registredExcludedKnm)]
    registredExcludedKnm.insert(0, ['', 'Всего', *[val for val in totalExcluded.values()]])
    registredExcludedKnm.insert(0, excludedReportTitle)
    title = []
    title.append('ТУ')
    for s in statusses:
        title.append(s)
        title.append('прирост')
    title.append('Всего на согласовании')
    xl.writeResultsInXL(results=value_list, title=title, pathFile=pathFile, sheetTitle='Общий отчет')
    xl.writeResultsInXL(results=registredExcludedKnm,
                        title=f'Регионы, в которых зарегистрированы исключения КНМ органами прокуратуры на {today}',
                        pathFile=pathFile, sheetIndex=1, sheetTitle='Исключенные')
    with open(previewsResultFile, 'w', encoding='utf-8') as file:
        json.dump(results, file)
    dailyAggreedProcessStyles(pathFile)


def reportDayliAggreedProcess(count_attribute='knm'):
    from datetime import datetime
    from private_config import dayliProcessFile
    import os
    today = str(datetime.now().strftime('%d.%m.%Y'))
    reportByAggreedProcess(dayliProcessFile(f"ежедневный {today} {count_attribute}.xlsx"), today,
                           count_by=count_attribute)
    isklListPath = dayliProcessFile(f"Исключения {today}.xlsx")
    if not os.path.exists(isklListPath):
        reportByProsecutorComments(isklListPath)
    return 'Отчет готов!'


def findIsklInfo(dirPath):
    import openpyxl
    wb = openpyxl.load_workbook(dirPath)
    sh = wb['причины исключений']
    tu = sh['A1'].value
    iskl_list = []
    for row in sh.iter_rows(min_row=3, values_only=True):
        iskl_list.append([tu, row[1], row[2]])
    return iskl_list


def mergeIskl(pathDir):
    import os
    import xl
    from private_config import dayliProcessFile
    from tqdm import tqdm

    iskl_list = []
    for file in tqdm(os.listdir(pathDir)):
        dirPath = os.path.join(pathDir, file)
        iskl_list.extend(findIsklInfo(dirPath))
    newFile = dayliProcessFile('все причины исключений.xlsx')
    xl.writeResultsInXL(results=iskl_list, pathFile=newFile, title='Совокупность исключений')


def useDatabase():
    from sql import Database
    d = Database(init_erknm=False)
    d.exec_it_database()
    return 'Работа с базой данных завершена!'


def groupByRegions(pathFile):
    import openpyxl
    import xl
    import os
    from datetime import datetime
    from pathlib import Path

    t_date = datetime.now().strftime('%d.%m.%Y')

    parentDir = Path(pathFile).parent

    todayDir = os.path.join(parentDir, f'на {t_date}')
    isklAnalysDir = os.path.join(todayDir, 'анализ исключений')
    if not os.path.exists(todayDir):
        os.mkdir(todayDir)
    if not os.path.exists(isklAnalysDir):
        os.mkdir(isklAnalysDir)

    datas = {}
    wb = openpyxl.load_workbook(pathFile)
    sh = wb.worksheets[0]
    for row in sh.iter_rows(min_row=2, values_only=True):
        tu = row[1]
        if row[1] in datas:
            datas[tu].append(row)
        else:
            datas[tu] = [row]
    for region, rows in datas.items():
        xl.writeResultsInXL([['', 'Причины', 'Встречается в КНМ'], *rows], title=region,
                            pathFile=os.path.join(isklAnalysDir, f'{region}.xlsx'), sheetTitle='Исключения')


def reportByIsklReasons(pathDir):
    import os
    import xl
    from fts import Fts
    from pathlib import Path
    from tqdm import tqdm
    from Dictionary import topIsklReasons
    from private_config import dayliProcessFile

    f = Fts()
    full_dict = {}
    value_list = []
    for file in tqdm(os.listdir(pathDir)):

        if 'xlsx' in file:
            tuName = Path(file).stem
            print(tuName)
            pathFile = os.path.join(pathDir, file)
            cols = xl.bringCol(pathFile, minRow=3, minCol=2, sheetIndex=1, colNumber='0:3')
            rows = dict([reas for reas in zip(cols[0], cols[1])])

            full_dict[tuName] = f.mapInTop(rows)

    # pprint(dict(sorted(full_dict.items(), key=lambda x: sum(x[1].values()), reverse=True)))
    topValues = topIsklReasons.values()

    for tu, value in full_dict.items():
        value_list.append([tu, *[full_dict[tu][val] for val in topValues]])

    xl.writeResultsInXL(results=value_list, title=['', *list(topValues)],
                        pathFile=dayliProcessFile("все.xlsx"))


def reportByIndicators():
    from private_config import dayliProcessFile
    import xl
    from mongo_database import WorkMongo
    wm = WorkMongo()
    knms = convertForsaving(wm.reportByRiskIndicatorKNM())
    pathFile = dayliProcessFile('выгрузка по индикаторным проверкам.xlsx')
    xl.writeResultsInXL(results=knms[1], title=knms[0], pathFile=pathFile)

    print(pathFile)
    return f'Выгрузка по индикаторам риска готова!'


def countIsklByReasons(pathDir):
    import xl
    from Nat import get_reasons, get_reasons_multy
    import os
    from pathlib import Path

    for file in os.listdir(pathDir):
        print(file)
        pathFile = os.path.join(pathDir, file)
        comments = xl.bringCol(pathFile, minRow=3)
        categories = get_reasons(comments)
        results = [['', 'причины исключений', 'Найдено в КНМ'], *sorted(
            [[n + 1, str(value['explanation']), int(value['count'])] for n, value in enumerate(categories.values())],
            key=lambda x: x[2], reverse=True)]
        # pprint(results)
        xl.writeResultsInXL(results=results, title=Path(pathFile).stem, pathFile=pathFile, sheetIndex=1,
                            sheetTitle='причины исключений')
    return 'Причины исключений собраны!'


def reportByVk():
    import xl
    from mongo_database import WorkMongo
    from private_config import dayliProcessFile
    from datetime import datetime

    now = datetime.now().strftime('%d.%m.%Y %H.%M')

    wm = WorkMongo()
    vkKinds = wm.reportByProductionConsist()
    finalList = []

    allProductionKinds = []
    results = {}
    # results = {'tu': {'prod1': 'count', 'prod2': 'count'}}

    for vkkind in vkKinds:
        tu = vkkind['_id']['tu']
        production = vkkind['_id']['production']
        totalCount = vkkind['totalCount']
        if not production in allProductionKinds:
            allProductionKinds.append(production)
        if tu in results:
            results[tu][production] = totalCount
        else:
            results[tu] = {production: totalCount}

    for n, (tu, productions) in enumerate(results.items()):
        finalList.append([n+1, tu, *[productions[kind] if kind in productions else 0 for kind in allProductionKinds]])

    pathFile = dayliProcessFile(f'выгрузка по составу согласованных объектов выборочного контроля {now}.xlsx')
    xl.writeResultsInXL(results=finalList, title=['', '', *allProductionKinds], pathFile=pathFile)
    print(len(allProductionKinds))
    return 'Выгрузка по составу согласованных объектов выборочного контроля готова!'


def getKnmWithout336PP():
    import xl
    from mongo_database import WorkMongo
    from private_config import dayliProcessFile

    wm = WorkMongo()

    knms = wm.getKNMWithoutBudjets()
    print(list(knms))
    results = convertForsaving(list(knms))

    title = results[0]
    results = [[r[0]['tu'], r[1]] for r in results[1]]

    # print(results[])
    print(results)
    xl.writeResultsInXL(title=title, results=results)


def downdoalZppInspection():
    import xl
    from mongo_database import WorkMongo
    from Dictionary import erknmFields


    wm = WorkMongo()
    knms = wm.getZPPInspections()
    new_knms = []
    for knm in knms:
        reason = knm['reasonsList'][0]['text'] if knm['reasonsList'] else 'Не заполнен'
        knm['reasonsList'] = reason
        new_knms.append(knm)
    # pprint(f"{list(knms)[0]=}")

    results = convertForsaving(list(new_knms))

    titles = []
    for field in results[0]:
        if field in erknmFields:
            titles.append(erknmFields[field])
        else:
            titles.append(field)
    #
    xl.writeResultsInXL(title=results[0], results=results[1])


def getObjectFromRHSByTu():
    import xl
    from mongo_database import WorkMongo
    from Dictionary import public_catering_kind

    wm = WorkMongo('rhs')

    _objects = wm.getRhsObjectsByTu(
        objectsKind=public_catering_kind,
        objectsRisk=['Высокий риск']
    )
    # pprint(list(_objects))
    result = convertForsaving(list(_objects))
    xl.writeResultsInXL(title=result[0], results=result[1])


def getObjectsFromRHSByTuRisk():
    import xl
    from mongo_database import WorkMongo
    import Dictionary as d

    wm = WorkMongo('rhs')


    #  количество категорий риска по ТУ, сопоставимых по видам деятельности с 1 индикатором ()
    tuRisksInfectIndicator = wm.getRhsObjectsTuRiskKind(
        kinds=[
            *group_kinds['Деятельность в сфере здравоохранения'],
            *group_kinds['Детские лагеря с дневным пребыванием'],
            *group_kinds['Деятельность общеобразовательных организаций'],
            *group_kinds['Деятельность дошкольных образовательных организаций'],
            *group_kinds['Объекты промышленности и транспорта'],
            *group_kinds['Деятельность водоснабжения и водоотведения'],
            *group_kinds['Коммунальное обслуживание'],
            *group_kinds['Деятельность по предоставлению персональных услуг'],
            *group_kinds['Деятельность гостиниц и прочих мест для временного проживания'],
            *group_kinds['Предоставление социальных услуг'],
            *group_kinds[
                'Деятельность детских и подростковых организаций, образования, в том числе дополнительного образования'],
            *group_kinds['Деятельность в области обращения с отходами'],
            *group_kinds['Работы с микроорганизмами'],
            *group_kinds['Организации для детей-сирот'],
            *group_kinds['Деятельность по организации отдыха и развлечений, культуры и спорта'],
            *group_kinds['Профессиональные образовательные организации'],
            *group_kinds['Учреждения высшего профессионального образования'],
        ]
    )
    unpackedTuRisksInfectIndicator = unpac_idAggregation(list(tuRisksInfectIndicator))
    resultTuRisksInfectIndicator = convertForsaving(unpackedTuRisksInfectIndicator)
    xl.writeResultsInXL(
        title=resultTuRisksInfectIndicator[0],
        results=resultTuRisksInfectIndicator[1],
        sheetIndex=0,
        sheetTitle="для индикатора 1",
        pathFile=os.path.join(default_path_to_save_result, "Категории риска по ТУ для расчета доли нереализованного риска.xlsx")
    )

    #  количество категорий риска по ТУ, сопоставимых по видам деятельности со 2 индикатором ()
    tuRisksParasitIndicator = wm.getRhsObjectsTuRiskKind(
        kinds=[
            *group_kinds['Деятельность в сфере здравоохранения'],
            *group_kinds['Детские лагеря с дневным пребыванием'],
            *group_kinds['Деятельность общеобразовательных организаций'],
            *group_kinds['Деятельность дошкольных образовательных организаций'],
            *group_kinds['Объекты промышленности и транспорта'],
            *group_kinds['Деятельность водоснабжения и водоотведения'],
            *group_kinds['Коммунальное обслуживание'],
            *group_kinds['Деятельность по предоставлению персональных услуг'],
            *group_kinds['Деятельность гостиниц и прочих мест для временного проживания'],
            *group_kinds['Предоставление социальных услуг'],
            *group_kinds['Деятельность детских и подростковых организаций, образования, в том числе дополнительного образования'],
            *group_kinds['Деятельность в области обращения с отходами'],
            *group_kinds['Работы с микроорганизмами'],
            *group_kinds['Организации для детей-сирот'],
            *group_kinds['Деятельность по организации отдыха и развлечений, культуры и спорта'],
            *group_kinds['Профессиональные образовательные организации'],
            *group_kinds['Учреждения высшего профессионального образования'],
            *group_kinds['Деятельность организаторов детского питания'],
            *group_kinds['Торговля пищевыми продуктами'],
            *group_kinds['Производство пищевых продуктов'],
            *group_kinds['Общественное питание населения'],
            *group_kinds['Продукция'],
        ]
    )
    unpackedTuRisksParasitIndicator = unpac_idAggregation(list(tuRisksParasitIndicator))
    resultTuRisksParasitIndicator = convertForsaving(unpackedTuRisksParasitIndicator)
    xl.writeResultsInXL(
        title=resultTuRisksParasitIndicator[0],
        results=resultTuRisksParasitIndicator[1],
        sheetIndex=1,
        sheetTitle="для индикатора 2",
        pathFile=os.path.join(default_path_to_save_result,
                              "Категории риска по ТУ для расчета доли нереализованного риска.xlsx")
    )

    #  количество категорий риска по ТУ, сопоставимых по видам деятельности с 3 индикатором ()
    tuRisksOKIIndicator = wm.getRhsObjectsTuRiskKind(
        kinds=[
            *group_kinds['Деятельность организаторов детского питания'],
            *group_kinds['Торговля пищевыми продуктами'],
            *group_kinds['Производство пищевых продуктов'],
            *group_kinds['Общественное питание населения'],
        ]
    )
    unpackedTuRisksOKIIndicator = unpac_idAggregation(list(tuRisksOKIIndicator))
    resultTuRisksOKIIndicator = convertForsaving(unpackedTuRisksOKIIndicator)
    xl.writeResultsInXL(
        title=resultTuRisksOKIIndicator[0],
        results=resultTuRisksOKIIndicator[1],
        sheetIndex=2,
        sheetTitle="для индикатора 3",
        pathFile=os.path.join(default_path_to_save_result,
                              "Категории риска по ТУ для расчета доли нереализованного риска.xlsx")
    )

    #  количество категорий риска по ТУ, сопоставимых по видам деятельности с 3 индикатором только продукция
    tuRisksOKIIndicatorProduction = wm.getRhsObjectsTuRiskKind(
        kinds=[
            *group_kinds['Продукция'],
        ]
    )
    unpackedTuRisksOKIIndicatorProduction = unpac_idAggregation(list(tuRisksOKIIndicatorProduction))
    resultTuRisksOKIIndicatorProduction = convertForsaving(unpackedTuRisksOKIIndicatorProduction)
    xl.writeResultsInXL(
        title=resultTuRisksOKIIndicatorProduction[0],
        results=resultTuRisksOKIIndicatorProduction[1],
        sheetIndex=3,
        sheetTitle="для индикатора 3П",
        pathFile=os.path.join(default_path_to_save_result,
                              "Категории риска по ТУ для расчета доли нереализованного риска.xlsx")
    )





def downloadForInspect():
    import os
    import xl
    from mongo_database import WorkMongo
    from Dictionary import tuCodeRegion, getActualTuName
    import json

    wm = WorkMongo('knm')
    knms = wm.reportForInspectSite()
    knms1 = []

    for knm in knms:
        durationHours = int(knm['durationHours']) if knm['durationHours'] else 0
        if 'Малое предприятие' in knm['mspCategory']:
            msp = 1
        elif 'Микропредприятие' in knm['mspCategory']:
            msp = 2
        else:
            msp = None

        if not msp:
            durationDays = knm['directDurationDays'] if knm['directDurationDays'] else knm['durationDays'] if knm['durationDays'] else 10
            durationDays = int(durationDays)
            durationDays = durationDays if durationDays <= 10 else 10
        else:
            durationDays = None
            durationHours = durationHours if durationHours else 50 if msp == 1 else 15

        places = '; '.join(knm['addresses']).replace('\n', ' ').replace('\t', ' ').strip()
        all_duration = " ".join([f"{durationDays if durationDays >= 10 else 10} дней" if durationDays else "", f"{durationHours} часов" if durationHours else ""]).strip()
        allCollaboratingOrganizations = '; '.join(knm['collaboratingOrganizations']).replace('\n', ' ').replace('\t', ' ').strip()
        controllingOrganization = getActualTuName(knm['controllingOrganization'])
        codeRegion = tuCodeRegion[controllingOrganization]

        knms1.append({
            'codeRegion': codeRegion,
            'controllingOrganization': controllingOrganization,
            'organizationName': knm['organizationName'],
            'ogrn': knm['ogrn'],
            'inn': knm['inn'],
            'addresses': places,
            'startDate': knm['startDate'],
            'duration': all_duration,
            'collaboratingOrganizations': allCollaboratingOrganizations,
            'kind': knm['kind'],
            'status': "Исключена" if knm['status'] == "Исключена" else ''
            # 'id': knm['id'],
            # 'erpId': knm['erpId'],
            # 'mspCategory': 'Микропредприятие' if msp == 2 else 'Малое предприятие' if msp == 1 else "Не является субъектом МСП",
        })
    # print(knms1)
    results = convertForsaving(knms1)
    saveDirPath = "C:\\Users\zaitsev_ad\Desktop"
    xl.writeResultsInXL(results=results[1], title=results[0], pathFile=os.path.join(saveDirPath, "кнм для размещения на сайте.xlsx"))
    with open(os.path.join(saveDirPath, "кнм для размещения на сайте.json"), "w", encoding="UTF-8") as file:
        json.dump(knms1, file)



def reportKnmByDates(year):
    from Dates_manager import getListDaysFromYear, split_year_for_periods
    from merge_tu import getActualName
    from mongo_database import WorkMongo
    from Dictionary import tuList
    import xl

    allDaysInYear = getListDaysFromYear(int(year))

    wm = WorkMongo('knm')
    tu_dict = {}
    title = ['ТУ']
    result = []
    for tu in tuList:
        tu_dict[tu] = {day: 0 for day in allDaysInYear}

    for date in tqdm(allDaysInYear, desc='Обработка запроса по каждому дню'):
        title.append(date)
        knm_tu = wm.getKnmFromDate(date)
        for knm in knm_tu:

            tu_dict[getActualName(knm["_id"])][date] += knm['count']

    for tu, days in tu_dict.items():

        for knmCount in days.values():
            pass
        result.append([tu, *[knmCount for knmCount in days.values()]])

    xl.writeResultsInXL(title=title, results=result, pathFile="C:\\Users\zaitsev_ad\Desktop\кнм по дням.xlsx")

    return 'Завершено!'

    # for day in allDaysInYear:
    #     knmsTu = wm.getKnmFromDate(day)


def load_rhs():
    from EIAS import Eias
    while True:
        if Eias().loadAllObjects():
            break


def loadEffectiveIndicatorsOfTu(path):
    import openpyxl
    from sql import Database
    wb = openpyxl.load_workbook(path)
    sh = wb.worksheets[0]

    results = []
    # print(results).

    for row in sh.iter_rows(min_row=4, min_col=3, values_only=True):
        r = [str(d) for d in row]
        results.append(r)

    Database().loadDataToEffIndic(results)
    return 'Загрузка успешно завершена!'




if __name__ == '__main__':
    functions = {
        'count_iskl': {'action': countIsklByReasons,
                       'desc': 'Делает отчет, по причинам исключений проверок прокуратурой',
                       'args': ["путь до папки Анализ исключений"]},
        'group_iskl_by_regions': {'action': groupByRegions,
                                  'desc': 'Сортирует нарушения по регионам',
                                  'args': ["путь до файла"]},
        'report_by_reasons_iskl': {'action': reportByIsklReasons,
                                   'desc': 'Сводит в таблицу анализ исключений по регионам',
                                   'args': ["путь до папки Анализ исключений"]},
        'load_knm': {'action': load_knm, 'desc': 'загружает проверки из ЕРКМН',
                     'args': ["Год, за который нужно выгрузить проверки"]},
        'load_pm': {'action': load_pm, 'desc': 'загружает профилактические мероприятия из ЕРКМН',
                    'args': ["Год, за который нужно выгрузить профилактические мероприятия"]},
        'load_rhs': {'action': load_rhs, 'desc': 'загружает все объекты из РХС',
                    'args': []},
        'merge_iskl_reason': {'action': mergeIskl,
                              'desc': 'сводит в одну таблицу результаты после анализа причин исключений',
                              'args': [
                                  "путь до папки, где хранятся результаты анализа по причинам исключений (у файлов на листе 'причины исключений' должен быть анализ нарушение - количество)"]},
        'merge_tu': {'action': mergeTu,
                     'desc': 'забирает столбец из файла exel c ТУ и по регурялке проверяет, кто есть в списке и сколько раз, а кого нет',
                     'args': ["путь до файла с ТУ в столбце А", 'номер столбца, который берем']},
        'report_by_diened_objects': {'action': reportByDienedObjects,
                                     'desc': 'отчет по объектам, исключенным в ходе проверки плана прокуратурой',
                                     'args': ['аттрибут, по которому будем считать: "knm" или "objects"']},
        'report_by_accept_objects': {'action': reportByAccepedObjects,
                                     'desc': 'отчет по объектам, исключенным в ходе проверки плана прокуратурой',
                                     'args': ['аттрибут, по которому будем считать: "knm" или "objects"']},
        'report_by_objects': {'action': reportByObjects, 'desc': 'делает отчет о количестве видов деятельности',
                              'args': ["Путь до файла, куда вносятся данные/если такого файла нет-создадим",
                                       """фильтры для поиска, в формате "{'k': 'v'}"""]},
        'report_by_aggreed': {'action': reportByAggreedProcess, 'desc': 'делает отчет о ходе согласования по регионам',
                              'args': ["Путь до файла, куда вносятся данные/если такого файла нет-создадим"]},
        'report_daily_aggreed': {'action': reportDayliAggreedProcess,
                                 'desc': 'делает ежедневный отчет о ходе согласования по регионам в папке C:\\Users\zaitsev_ad\Documents\ЕРКНМ\План 2024\этап планирования\ежедневный мониторинг процесса согласования',
                                 'args': ['аттрибут, по которому будем считать: "knm" или "objects"']},
        'report_by_risks': {'action': reportByRisks, 'desc': 'делает отчет о количестве категорий риска объектов',
                            'args': ["Путь до файла, куда вносятся данные/если такого файла нет-создадим",
                                     """фильтры для поиска, в формате "{'k': 'v'}"""]},
        'report_by_consists_vk': {'action': reportByVk,
                                  'desc': 'делает отчет о составе продукции выборочного контроля',
                                  'args': []},
        'report_by_risksIndicators': {'action': reportByIndicators,
                                      'desc': 'делает выгрузку по проверкам, основаниями для которых стали срабатывания индикаторов риска',
                                      'args': []},
        'report_by_fk': {'action': reportByFreeCommand, 'desc': 'выгружает отчет по введенной команде',
                         'args': ["Путь до файла, куда вносятся данные/если такого файла нет-создадим",
                                  """команда для поиска, в формате "{'k': 'v'}"""]},
        'report_by_iskl': {'action': reportByProsecutorComments,
                           'desc': 'выгружает причины исключений с номерами проверок',
                           'args': ["Путь до файла, куда вносятся данные/если такого файла нет-создадим"]},
        'report_without_336': {'action': getKnmWithout336PP, 'desc': 'делает выгрузку по регионам исключенных 336 ПП', 'args': []},
        'downdoal_zpp_inspection': {'action': downdoalZppInspection, 'desc': 'Выгрузка для Управления мониторинга за продукцией в обороте', 'args': []},
        'report_rhs_by_kind_object': {'action': getObjectFromRHSByTu,
                                    'desc': 'Отчет по составу определенных объектов в РХС по ТУ', 'args': []},
        'report_knm_for_inspectrospotrebnadzor': {'action': downloadForInspect,
                                      'desc': 'выгрузка КНМ для загрузки на сайт Роспотребнадзора inspect.rospotrebnadzor.ru', 'args': []},
        'report_knm_byDates': {'action': reportKnmByDates,
                                                  'desc': 'выгрузка КНМ по каждому дню в году в разрезе ТУ',
                                                  'args': ["Год КНМ"]},
        'report_for_unrealized_risk': {'action': getObjectsFromRHSByTuRisk, 'desc': 'Делает выгрузку по категориям риска в ТУ для расчета нереализованного риска', 'args': []},
        'load_tu_sql_tu_indicators': {'action': loadEffectiveIndicatorsOfTu, 'desc': 'Загружает в базу данных кнд sql индикаторы эффективности из файла', 'args': ["Путь к файлу показатели эффективности"]},
        'use_database': {'action': useDatabase, 'desc': 'Дает интерактивный доступ в базу данных doc/knd', 'args': []},

    }
    # getObjectsFromRHSByTuRisk
    arg1 = argv[1] if len(argv) >= 2 else None
    if arg1 in functions:
        # print([None if len(argv) < n+1 else 1 for n in range(functions[arg1]['args'])])
        args = [argv[n + 2] if len(argv) > n + 2 else None for n in range(len(functions[arg1]['args']))]
        if args:
            pprint(functions[arg1]['action'](*args))
        else:
            pprint(functions[arg1]['action']())

    if arg1 == 'help' or not arg1:
        print('------------------------------')
        for name, info in functions.items():
            lenName = len(name)
            middleTire = f"{' ' * (lenName + 2)}{'-' * (30 - lenName)}"
            print(f'{name}  |  {info["desc"]}')
            print(middleTire)
            print(f"{' ' * (lenName + 2)}Аргументы:")
            print(middleTire)

            for n, _arg in enumerate(info["args"]):
                print(f"{' ' * (lenName + 5)}{n + 1}.) {_arg}")
                print(middleTire)
            print('------------------------------')
