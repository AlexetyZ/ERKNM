import openpyxl

operativ_table_path = 'Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx'
main_table_path = 'Z:\\Для ВСЕХ\\ПЛАН 2023\\ПЛАН 2023 по МЕСЯЦАМ.xlsx'

operativ_table = openpyxl.load_workbook(operativ_table_path)
main_table = openpyxl.load_workbook(main_table_path)

ot_sh = operativ_table.worksheets[0]
mt_sh = main_table.worksheets[0]

for ot_cell in ot_sh['Z']:
    ot_knm_value = ot_sh.cell(row=ot_cell.row, column=19).value
    ot_id_knm = ot_cell.value

    if ot_cell.value is None or ot_knm_value is None:
        continue
    else:
        print(ot_knm_value)
        knm_number = ot_knm_value.split(' ')[1]

        for mt_cell in mt_sh['Z']:
            mt_id_knm = mt_cell.value
            if mt_id_knm == ot_id_knm:
                mt_sh.cell(row=mt_cell.row, column=9, value=knm_number)
                mt_sh.cell(row=mt_cell.row, column=12, value='эпид. отдел')

main_table.save(main_table_path)

