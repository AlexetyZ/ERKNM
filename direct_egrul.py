import os.path
import random
import time

from functools import partial

from EGRUL import Egrul
import openpyxl


class Direct:

    def __init__(self, interactive: bool = True):
        if interactive:
            navigation = ['1', '2']
            main_menu = input(
                'Для навигации в приложении нужно ввести соответствующую цифру из предложенных\nВыберите действие:\n 1. быстро найти организацию в егрюл\n 2. найти сведения об организациях по ИНН в файле XLSX\n')
            while True:

                if main_menu in navigation:
                    break
                else:
                    main_menu = input(
                        'Введено неверное значение. Для навигации в приложении нужно ввести соответствующую цифру из предложенных\nВыберите действие:\n 1. быстро найти организацию в егрюл\n 2. найти сведения об организациях по ИНН в файле XLSX\n')

            if main_menu == '1':
                while True:
                    org_info = self.get_org_info()
                    if org_info == 'q':
                        break
                    result = self.find_full_info(org_info)
                    print('')
                    for key, value in result.items():
                        print(f'{key} - {value}')
                    command = input('\nПри копировании информации, сочетание клавиш CTRL+C без выделения фрагментов может привести к закрытию терминала\n\nДля завершения введите "q" или "й", для продолжения введите любое другое значение.\n')
                    if command == 'q' or command == 'й':
                        break
            if main_menu == '2':
                while True:
                    path = self.get_exelfile_path()
                    if os.path.exists(path):
                        break
                    else:
                        print("Такого файла не существует, введите корректный путь до существующего файла формата xlsx")

                inn_letter = input('Введите букву столбца, содержащий ИНН (большую латинскую)\n')
                ogrn_letter = input('Введите букву столбца, куда следует вставлять найденную информацию (большую латинскую)\n')
                self.get_values_from_table(path, inn_letter, ogrn_letter)
                time.sleep(5)

    def get_exelfile_path(self):
        print(
            'Абсолютный путь до файла - путь начитая с расположения на диске/папке/и тд. Его можно получить через буфер обмена методом "Скопировать путь"')
        print("ВАЖНО!!!")
        print()
        print(
            'Для корректной работы, введите столбец, откуда будем брать ИНН, а так же, столбец, куда будем вставлять Нужную информацию. Начало перебора ИНН из таблицы начнется со строки 2. Убедитесь, что файл не открыт и сохранен в корректном формате(".xlsx")')
        print()
        print("ВАЖНО!!!")
        paths = input('ВВЕДИТЕ АБСОЛЮТНЫЙ ПУТЬ ДО ОБРАБАТЫВАЕМОГО ФАЙЛА\n')
        path = paths.replace(paths[:1], '').replace(paths[:-1], '')
        return path

    def get_values_from_table(self, path, inn_letter, ogrn_letter):
        navigation = ['1', '2', '3', '4']
        choose_func = input('Выберете функционал поиска: \n1. ОГРН, \n2. Название, \n3. Адрес, \n4. Информация о ликвидиации\n')

        while True:

            if choose_func in navigation:
                break
            else:
                choose_func = input('Введенное значение выходит за пределы навигации. \nВыберете функционал поиска: \n1. ОГРН, \n2. Название, \n3. Адрес, \n4. Информация о ликвидиации\n')

        wb = openpyxl.load_workbook(path)
        sh = wb.active
        inn_table = sh[f'{inn_letter}']
        ogrn_table = sh[f'{ogrn_letter}']
        for inn_cell in inn_table:
            if inn_cell.row > 1:
                if inn_cell.value is None:
                    print('Пустой ИНН')
                    continue
                if sh.cell(inn_cell.row, column=ogrn_table[0].column).value is not None:
                    print('Уже заполнено огрн')
                    continue

                # print(inn_cell.value)
                inn = str(inn_cell.value)
                if choose_func == '1':
                    result = self.find_info(inn)
                elif choose_func == '2':
                    result = self.find_info_name(inn)
                elif choose_func == '3':
                    result = self.find_info_address(inn)

                elif choose_func == '4':
                    result = self.find_info_liquidate(inn)
                else:
                    print("не удалось распознать")
                    wb.save(path)
                    return 0

                # print(result)



                sh.cell(row=inn_cell.row, column=ogrn_table[0].column, value=result)
                wb.save(path)
                time.sleep(1)

    def get_org_info(self):
        org_info = input('введите известную информацию об учреждении (чем специфичнее информация, тем точнее результат, так что лучше вводить ИНН или ОГРН)\n')
        return org_info

    def find_full_info(self, org_info):
        find_in_egrul = Egrul().find_in_egrul
        try:
            request = find_in_egrul(org_info)
        except Exception as ex:
            raise ValueError(f'Ошибка запроса: {ex}')
        result = {}
        try:
            name = request['n']
        except:
            result["Ошибка"] = "Не найдено ни одного результата с таким запросом"
            return result
        inn = request['i']
        ogrn = request['o']
        try:
            address = request['a']
        except:
            address = 'Нет информации об адресе'

        try:
            liquidate_info = request['e']
        except:
            liquidate_info = 'Нет информации о ликвидации'
        result['Наименование'] = name

        result['ИНН'] = inn
        result['ОГРН'] = ogrn
        result['Адрес'] = address
        result['Информация о ликвидации'] = liquidate_info

        return result

    def find_info(self, org_info):
        find_in_egrul = Egrul().find_in_egrul
        try:
            request = find_in_egrul(org_info)
        except Exception as ex:
            raise ValueError(f'Ошибка запроса: {ex}')
        result = request['o']
        print(result)
        return result

    def find_info_name(self, org_info):
        find_in_egrul = Egrul().find_in_egrul
        try:
            request = find_in_egrul(org_info)
        except Exception as ex:
            raise ValueError(f'Ошибка запроса: {ex}')
        result = request['n']
        print(result)
        return result

    def find_info_address(self, org_info):
        find_in_egrul = Egrul().find_in_egrul
        try:
            request = find_in_egrul(org_info)
        except Exception as ex:
            raise ValueError(f'Ошибка запроса: {ex}')
        try:
            result = request['a']
        except:
            result = 'Нет адреса'
        print(result)
        return result

    def find_info_liquidate(self, org_info):
        find_in_egrul = Egrul().find_in_egrul
        try:
            request = find_in_egrul(org_info)
        except Exception as ex:
            raise ValueError(f'Ошибка запроса: {ex}')
        try:
            result = request['e']
        except:
            result = 'Информация о ликвидации отсутствует'
        print(result)
        return result


if __name__ == '__main__':
    Direct()

