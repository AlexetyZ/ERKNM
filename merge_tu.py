import openpyxl
from Dictionary import tuRegExp, tuList
import re


def main(file, tu_cell_number, cell_number_to_paste, sheetName=None, sheetIndex=None):
    wb = openpyxl.load_workbook(file)
    # if sheetName is None and sheetIndex is None:
    #     sh = wb.worksheets[0]
    # elif sheetName:
    #     sh = wb[sheetName]
    # else:
    #     sh = wb.worksheets[sheetIndex]

    if sheetIndex:
        sh = wb.worksheets[sheetIndex]
    elif sheetName:
        sh = wb[sheetName]
    else:
        sh = wb.worksheets[0]

    for row in sh.iter_rows(min_row=2, max_col=cell_number_to_paste+1):
        tuCurrent = row[tu_cell_number].value
        for reg in tuRegExp:
            if re.search(reg, str(tuCurrent).lower().strip()):
                row[cell_number_to_paste].value = tuRegExp[reg]
                break
    wb.save(file)


def get_tu_iso(file, tu_cell_number, cell_number_to_paste, sheetName=None, sheetIndex=None):
    wb = openpyxl.load_workbook(file)

    if sheetIndex:
        sh = wb.worksheets[sheetIndex]
    elif sheetName:
        sh = wb[sheetName]
    else:
        sh = wb.worksheets[0]


    wb.save(file)


def getActualName(text):
    for reg in tuRegExp:
        if re.search(reg, str(text).lower().strip()):
            return tuRegExp[reg]
    return None


if __name__ == '__main__':

    main(
        file="C:\\Users\zaitsev_ad\Desktop\Справка об эффективности КНМ и ПМ (1).12\СЗЗ услуги статистика.xlsx",
        tu_cell_number=0,
        cell_number_to_paste=10,
        sheetIndex=0
    )
    # print(getActualName('свердл'))

