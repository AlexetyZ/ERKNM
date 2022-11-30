import time

from ERKNM import erknm
import datetime
import logging
import openpyxl
import Normative_documentation
import Proverochnii_list
import clipboard



def exel_to_json_Objects():
    wb_path = "Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx"
    wb = openpyxl.load_workbook(wb_path)

    # юридические лица
    ul = wb.worksheets[0]

    # отражение нормативной документации
    doc_proverki_ND = wb.worksheets[1]
    viesd_proverki_ND = wb.worksheets[2]
    III_proverki_ND = wb.worksheets[3]

    lenBB = len(ul['B:B'])

    UL_with_objects = {}
    for i in range(4, lenBB + 1):
        if ul[f'S{i}'].value != None:
            continue

        objects = {}

        if ul[f'D{i}'].value != None:
            subject = ul[f'A{i}'].value + '|' + ul[f'D{i}'].value
            objects.update({ul[f'E{i}'].value: {
                'адрес': str.strip(str(ul[f'G{i}'].value)),
                'начало проверки': str(ul[f'B{i}'].value),
                'окончание проверки': str(ul[f'C{i}'].value),
                "ОГРН": str(ul[f'L{i}'].value),
                "деятельность": str(ul[f'N{i}'].value),
                'категория риска': str(ul[f'P{i}'].value),
                'класс опасности': str(ul[f'Q{i}'].value),
                'дата последнего планового кнм': str(ul[f'R{i}'].value),
                'строчка': i}})
            UL_with_objects[subject] = objects

        if ul[f'D{i}'].value == None:
            # print(len(UL_with_objects))
            # print(list(UL_with_objects.values())[len(UL_with_objects)-1])



            # objects.update(list(UL_with_objects.values())[len(UL_with_objects) - 1])



            try:
                objects.update(list(UL_with_objects.values())[len(UL_with_objects) - 1])
            except:
                ul.cell(row=i, column=19, value=ul[f'S{i-1}'].value)
                ul.cell(row=i, column=20, value=ul[f'T{i-1}'].value)
                wb.save(wb_path)
                try:
                    objects.update(list(UL_with_objects.values())[len(UL_with_objects) - 1])
                except Exception as ex:
                    print(ex)
            object_type = ul[f'E{i}'].value
            while True:
                if object_type in objects:
                    object_type = object_type+'1'
                    continue
                if object_type not in objects:
                    break

            #
            objects.update({object_type: {
                'адрес': str.strip(str(ul[f'G{i}'].value)),
                'начало проверки': str(ul[f'B{i}'].value),
                'окончание проверки': str(ul[f'C{i}'].value),
                "ОГРН": str(ul[f'L{i}'].value),
                "деятельность": str(ul[f'N{i}'].value),
                'категория риска': str(ul[f'P{i}'].value),
                'класс опасности': str(ul[f'Q{i}'].value),
                'дата последнего планового кнм': str(ul[f'R{i}'].value),
                'строчка': i}})
            UL_with_objects[list(UL_with_objects.keys())[len(UL_with_objects) - 1]] = objects
            pass
    return UL_with_objects

def main():
    object_kinds = {
        'дошкольные образовательные организации': '174',
        'общеобразовательные организации': '175',
        'Деятельность детских лагерей на время каникул': '252',
        'школы-интернаты, специальные (коррекционные) общеобразовательные организации': '182',
        'перинатальные центры, родильные дома, родильные отделения': '162',
        'Объекты здравоохранения': '0',
        'дома (интернаты) для лиц с физическими или умственными недостатками, в том числе геронтопсихиатрические центры, психоневрологические интернаты': '161'
    }

    objects_risk = {
        'чрезвычайно высокий риск': '0',
        'высокий риск': '1',
        'значительный риск': '2',
        'средний риск': '3',
        'умеренный риск': '4',
    }

    object_danger = {
        'Первый': '0',
        'Второй': '1',
        'Третий': '2',
        'Четвертый': '3',
    }

    try:
        knms = exel_to_json_Objects()
        print(knms)
    except:
        knms = exel_to_json_Objects()
        print(knms)

    clipboard.copy('some information')

    # for knm in knms.items():
    #     subject_data = list(knm[1].items())
    #
    #     for n, sd in enumerate(subject_data):
    #         if n != 0:
    #             print(sd[1]['ОГРН'])


        # print(f"номер кнм {knm[0]}  --->>> ")

    #     print(subject_data[0][1]['строчка'])
    #     wb = openpyxl.load_workbook("Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx")
    #     # юридические лица
    #     result = {'number': 'получилось', 'status': 'фортануло'}
    #     row = int(subject_data[0][1]['строчка'])
    #     ul = wb.worksheets[0]
    #     ul.cell(row=row, column=19, value=result['number'])
    #     ul.cell(row=row, column=20, value=result['status'])
    #
    #     wb.save("Z:\\ДЛЯ ЭПИДЕМИОЛОГОВ\\2023\\забивать на 2023.xlsx")


    # print(len(Normative_documentation._0))










if __name__ == '__main__':
    main()